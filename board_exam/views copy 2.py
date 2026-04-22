from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.views import View
from django.template.loader import render_to_string
from weasyprint import HTML
import random, io, zipfile, uuid, re
import xml.etree.ElementTree as ET
from scripts.check import detect_objects, sort_objects_by_distance, group_and_sequence
from django.http import JsonResponse
import numpy as np
import cv2, time, os, json, base64, traceback
from django.contrib.auth import logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from django.views.decorators.csrf import csrf_protect
from django.db import IntegrityError
from django.db.models import Q
from .forms import AnswerSheetForm, SignUpForm, EmailAuthenticationForm 
from itertools import zip_longest
from django.http import HttpResponseRedirect, HttpResponseForbidden
from django.urls import reverse
from django.utils import timezone
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.conf import settings
import PyPDF2
import fitz, gc
import pdfplumber, docx, string
from docx import Document
from PyPDF2 import PdfReader
from django.core.files import File
import pandas as pd
from .config import BOARD_EXAM_TOPICS, LEVELS
from django.db import models, transaction
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.db.models import Count, Q, Avg, F
from collections import defaultdict
from django.views.decorators.http import require_http_methods
import datetime
import openai
from datetime import datetime
from scripts.model_loader import get_original_model, get_cropped_model
from .services.user_service import UserService
from .services.question_service import QuestionService
from .services.test_service import TestService
from .services.result_service import ResultService
from .services.practice_service import PracticeService
from google.cloud import storage
from firebase_admin import firestore, auth, storage
import firebase_admin
from functools import wraps
from .firebase import initialize_firebase


initialize_firebase()

def firebase_login_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        uid = request.session.get("uid")

        if not uid:
            return redirect("login")

        user = UserService.get_user(uid)

        # ✅ prevent crash early
        if not user:
            return redirect("login")

        request.user_data = user  # optional (cleaner access)

        return view_func(request, *args, **kwargs)
    return wrapper

logo_path = os.path.join(settings.BASE_DIR, 'static', 'EXIM2.png')  # full path


# =========================
# FIRESTORE TEST
# =========================
def test_firestore(request):
    doc_id = UserService.test_firestore()
    return JsonResponse({
        "ok": True,
        "doc_id": doc_id
    })

import logging

logger = logging.getLogger(__name__)

@csrf_protect
def signup(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)

        logger.info("SIGNUP POST HIT")

        if not form.is_valid():
            logger.warning(f"Form invalid: {form.errors}")
            return render(request, 'signup.html', {'form': form})

        try:
            role = form.cleaned_data['role']
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']

            logger.info(f"Creating Firebase user: {email}")

            user_record = auth.create_user(
                email=email,
                password=password
            )

            uid = user_record.uid

            logger.info(f"Firebase user created: {uid}")

            # ✅ PASS ALL EXTRA FIELDS (IMPORTANT)
            UserService.create_user(
                user_id=uid,
                email=email,
                role=role,
                is_student=(role == "student"),
                is_staff=(role == "teacher"),

                # 🔥 student + profile fields
                course=form.cleaned_data.get("course"),
                last_name=form.cleaned_data.get("last_name"),
                first_name=form.cleaned_data.get("first_name"),
                middle_name=form.cleaned_data.get("middle_name"),
                birthdate=str(form.cleaned_data.get("birthdate")) if form.cleaned_data.get("birthdate") else None,
            )

            logger.info("Firestore user created")

            messages.success(request, "Account created successfully!")
            return redirect('login')

        except Exception as e:
            logger.error("SIGNUP ERROR", exc_info=True)
            return HttpResponse(f"ERROR: {str(e)}")

    return render(request, 'signup.html', {'form': SignUpForm()})

# =========================
# LOGIN (FIREBASE TOKEN)
# =========================
import json

def login_view(request):
    if request.method == 'POST':
        try:
            logger.info("🔥 LOGIN START")

            data = json.loads(request.body.decode("utf-8"))
            logger.info(f"🔥 PARSED DATA: {data}")

            id_token = data.get("id_token")
            if not id_token:
                return JsonResponse({"error": "No ID token provided"}, status=400)

            decoded = auth.verify_id_token(id_token)
            uid = decoded["uid"]

            user_data = UserService.get_user(uid)

            if not user_data:
                return JsonResponse({"error": "User not found"}, status=404)

            role = user_data.get("role", "student")

            request.session["uid"] = uid
            request.session["role"] = role
            request.session.modified = True

            logger.info("🔥 LOGIN SUCCESS")

            return JsonResponse({
                "success": True,
                "redirect": "/home/" if role == "teacher" else "/home_student/"
            })

        except Exception as e:
            logger.error("🔥 LOGIN ERROR")
            logger.error(traceback.format_exc())

            # 🔥 ALWAYS RETURN JSON (NEVER HTML)
            return JsonResponse({
                "success": False,
                "error": str(e),
                "type": type(e).__name__
            }, status=500)

    return render(request, "login.html")


# =========================
# LOGOUT
# =========================
def logout_view(request):
    request.session.flush()
    return redirect('login')


# =========================
# ROLE HELPERS
# =========================
def get_user_role(request):
    uid = request.session.get("uid")
    if not uid:
        return None

    user_data = UserService.get_user(uid)
    return user_data.get("role") if user_data else None


# =========================
# DASHBOARD ROUTING
# =========================
@firebase_login_required
def main_dashboard(request):
    user = UserService.get_user(request.session["uid"]) 
    role = user.get("role")

    if role == "teacher":
        return redirect("home")
    elif role == "student":
        return redirect("home_student")
    else:
        return redirect("login")


# =========================
# TEACHER DASHBOARD
# =========================
def home(request):
    try:
        logger.warning("🔥 ENTER HOME VIEW")

        uid = request.session.get("uid")
        logger.warning(f"🔥 SESSION UID: {uid}")

        if not uid:
            logger.warning("❌ NO UID IN SESSION")
            return redirect("login")

        user = UserService.get_user(uid)
        logger.warning(f"🔥 FIRESTORE USER: {user}")

        if not user:
            logger.warning("❌ USER NOT FOUND IN FIRESTORE")
            return redirect("login")

        role = user.get("role")
        logger.warning(f"🔥 ROLE: {role}")

        if role != "teacher":
            return HttpResponseForbidden("You cannot access this page")

        return render(request, "home.html")

    except Exception as e:
        logger.error("🔥 HOME CRASH ERROR", exc_info=True)
        return JsonResponse({
            "error": str(e),
            "type": type(e).__name__
        }, status=500)


# =========================
# STUDENT DASHBOARD
# =========================
def home_student(request):
    uid = request.session.get("uid")

    if not uid:
        return redirect("login")

    user = UserService.get_user(uid)

    # ✅ PREVENT CRASH
    if not user:
        return redirect("login")

    role = user.get("role")

    if role != "student":
        return HttpResponseForbidden("You cannot access this page")

    return render(request, 'home_student.html')

# =========================
# ROOT REDIRECT
# =========================
def root_redirect(request):
    if not request.session.get("uid"):
        return redirect("login")

    user = UserService.get_user(request.session["uid"]) 
    role = user.get("role")

    if role == "teacher":
        return redirect("home")
    elif role == "student":
        return redirect("home_student")
    else:
        return redirect("login")



bucket = storage.bucket()

def serve_image(request, image_name):
    try:
        blob = bucket.blob(image_name)

        if not blob.exists():
            raise Http404("Image not found")

        return HttpResponse(
            blob.download_as_bytes(),
            content_type=blob.content_type or "image/jpeg"
        )

    except Exception:
        raise Http404("Image not found")


from urllib.parse import quote

def firebase_image_url(path):
    if not path:
        return None

    bucket = "project-5e6fa15a-0ef4-476a-b87.firebasestorage.app"

    return (
        "https://firebasestorage.googleapis.com/v0/b/"
        f"{bucket}/o/{quote(path, safe='')}"
        "?alt=media"
    )

def question_bank(request):

    questions = QuestionService.get_all()

    letters = ["A", "B", "C", "D", "E"]

    for q in questions:
        choices = q.get("choices", [])

        q["lettered_choices"] = [
            {"letter": letters[i], "text": c.get("text")}
            for i, c in enumerate(choices[:5])
        ]

        correct_letter = q.get("correct_letter")

        # ✅ GET CORRECT ANSWER TEXT FROM LETTER
        correct_text = "-"
        if correct_letter:
            index = ord(correct_letter) - ord("A")  # A->0, B->1, etc.
            if 0 <= index < len(choices):
                correct_text = choices[index].get("text", "-")

        q["correct_answer_text"] = correct_text

        # formatting
        q["board_exam_names"] = ", ".join(q.get("board_exams", []))
        q["subject_names"] = ", ".join(q.get("subjects", []))
        q["topic_name"] = q.get("topic", "-")
        q["level_name"] = q.get("difficulty", "-")
        q["image_url"] = firebase_image_url(q.get("image"))

    return render(request, "question_bank.html", {"questions": questions})

class Add_Question(View):

    def get(self, request):
        return render(request, "add_question.html", {
            'BOARD_EXAMS': list(BOARD_EXAM_TOPICS.keys()),
            'LEVELS_JSON': json.dumps(LEVELS),
            'BOARD_EXAM_TOPICS_JSON': json.dumps(BOARD_EXAM_TOPICS),
        })

    def post(self, request):

        num_questions = sum(
            1 for key in request.POST.keys()
            if key.startswith("question_text_")
        )

        if num_questions == 0:
            messages.error(request, "No questions to save!")
            return redirect("add_question")

        for i in range(1, num_questions + 1):

            question_text = request.POST.get(f"question_text_{i}")
            if not question_text:
                continue

            board_exam_names = request.POST.getlist("board_exam_checkbox")
            subject_names = request.POST.getlist(f"subjects_{i}[]")
            topic_name = request.POST.get(f"topic_{i}")
            level_name = request.POST.get(f"level_{i}")
            source = request.POST.get(f"source_{i}", "google.com")

            # -------------------------
            # FIREBASE IMAGE UPLOAD (UBLA SAFE)
            # -------------------------
            image_file = request.FILES.get(f"image_{i}")
            image_path = None

            if image_file:
                # blob = bucket.blob(f"questions/{uuid.uuid4().hex}.jpg")
                blob = bucket.blob(
                        f"questions/{topic_name}/{uuid.uuid4().hex}_{image_file.name}"
                    )
                blob.upload_from_file(image_file)

               
                # ✅ STORE ONLY PATH
                image_path = blob.name

            # -------------------------
            # CHOICES
            # -------------------------
            choices = []
            correct_letter = request.POST.get(f"correct_answer_{i}")

            for letter in ["A", "B", "C", "D", "E"]:
                text = request.POST.get(f"choice{letter}_{i}")
                if text:
                    choices.append({
                        "text": text,
                        "is_correct": letter == correct_letter
                    })

            # -------------------------
            # VALIDATION
            # -------------------------
            if len(choices) < 2:
                messages.error(request, f"Question {i} must have at least 2 choices")
                return redirect("add_question")

            if not topic_name:
                messages.error(request, f"Question {i} must have a topic")
                return redirect("add_question")

            # -------------------------
            # SAVE TO FIRESTORE / DB
            # -------------------------
            QuestionService.create_question(
                question_text=question_text,
                choices=choices,
                image_url=image_path,   # IMPORTANT: this is now a PATH, not URL
                level=level_name,
                source=source,
                subject_names=subject_names,
                topic_name=topic_name,
                board_exam_list=board_exam_names
            )

        messages.success(request, f"{num_questions} questions added successfully!")
        return redirect("question_bank")


def reorder_choices(choices):
    special = []
    normal = []

    for c in choices:

        # handle both dict and string formats
        if isinstance(c, dict):
            text = c.get("text") or c.get("choice") or ""
        else:
            text = str(c)

        if text.strip().lower() in ["none of the above", "all of the above"]:
            special.append(c)
        else:
            normal.append(c)

    return normal + special
    
def generate_test(request):
    if request.method == "POST":

        board_exam = request.POST.get("board_exam")
        subject = request.POST.get("subject", "")
        num_questions = int(request.POST.get("num_questions", 0))

        questions = QuestionService.get_all()

        if board_exam:
            questions = [
                q for q in questions
                if board_exam in (q.get("board_exams") or [])
            ]

        if subject:
            questions = [
                q for q in questions
                if subject in (q.get("subjects") or [])
            ]

        if len(questions) < num_questions:
            return render(request, "generate_test.html", {
                "BOARD_EXAMS": list(BOARD_EXAM_TOPICS.keys()),
                "SUBJECTS_JSON": json.dumps(BOARD_EXAM_TOPICS),
                "error_message": "Not enough questions"
            })

        selected = random.sample(questions, num_questions)

        for q in selected:
            q["image_url"] = firebase_image_url(q.get("image"))

            # reorder choices so "None/All of the above" is always last
            if "choices" in q and q["choices"]:
                q["choices"] = reorder_choices(q["choices"])

        # -----------------------------
        # CREATE ONLY 2 TEST SETS
        # -----------------------------

        set_a_id = uuid.uuid4().hex
        set_b_id = uuid.uuid4().hex


        return render(request, "generated_test.html", {
            "set_a_questions": selected,
            "set_b_questions": random.sample(selected, len(selected)),
            "board_exam": board_exam,
            "subject": subject,
            "set_a_id": set_a_id,
            "set_b_id": set_b_id,
        })

    return render(request, "generate_test.html", {
        "BOARD_EXAMS": list(BOARD_EXAM_TOPICS.keys()),
        "SUBJECTS_JSON": json.dumps(BOARD_EXAM_TOPICS)
    })
    

def map_letter_text(choices_lists, correct_text_dict):

    answer_key = {}
    num_choices = len(choices_lists)
    letters = list(string.ascii_uppercase[:num_choices])

    for i, correct_text in correct_text_dict.items():

        choice_map = {
            letters[idx]: choices_lists[idx][i-1]
            for idx in range(num_choices)
        }

        correct_letter = next(
            (l for l, t in choice_map.items() if t == correct_text),
            None
        )

        answer_key[str(i)] = {
            "letter": correct_letter,
            "text": correct_text
        }

    return answer_key

def get_questions_with_choices(question_docs):
    questions = []
    letters = ['A', 'B', 'C', 'D', 'E']

    for q in question_docs:
        # q is already a dict now
        choices = q.get("choices", [])

        formatted_choices = []
        for i, c in enumerate(choices[:5]):
            formatted_choices.append({
                "letter": letters[i],
                "text": c.get("text", "")
            })

        questions.append({
            "id": q.get("id"),
            "question": q.get("question_text"),
            "choices": formatted_choices,

            #  ED: your field is "image", not "images"
            "image": q.get("image")
        })

    return questions

def build_answer_key(question_docs):

    letters = ['A', 'B', 'C', 'D', 'E']
    answer_key = {}

    for i, q in enumerate(question_docs, start=1):

        choices = q.get("choices", [])

        for idx, c in enumerate(choices):
            if c.get("is_correct"):
                answer_key[str(i)] = {
                    "letter": letters[idx],
                    "text": c.get("text", "")
                }
                break

    return answer_key

def download_test_interface(request):
    test_keys = TestService.get_all_tests()

    test_keys = sorted(
        test_keys,
        key=lambda x: x.get("created_at").timestamp() if x.get("created_at") else 0,
        reverse=True
    )

    uid = request.session.get("uid")
    user = UserService.get_user(uid) if uid else None
    role = (user or {}).get("role")

    return render(request, 'download_test.html', {
        'test_keys': test_keys,
        'role': role
    })



def download_test_pdf(request):
    if request.method != "POST":
        return HttpResponse("Invalid request method", status=405)

    try:
        subject_name = request.POST.get("subject")
        board_exam_name = request.POST.get("board_exam")

        set_a_ids = request.POST.getlist("set_a_question_ids[]")
        set_b_ids = request.POST.getlist("set_b_question_ids[]")

        # -------------------------
        # FETCH VIA SERVICE (NO db)
        # -------------------------
        set_a_docs = [
            QuestionService.get(qid)
            for qid in set_a_ids
        ]
        set_b_docs = [
            QuestionService.get(qid)
            for qid in set_b_ids
        ]

        set_a_docs = [q for q in set_a_docs if q]
        set_b_docs = [q for q in set_b_docs if q]

        # -------------------------
        # NORMALIZE IMAGE URL
        # -------------------------
        for q in set_a_docs:
            q["image"] = firebase_image_url(q.get("image"))

        for q in set_b_docs:
            q["image"] = firebase_image_url(q.get("image"))

        questions_set_a = get_questions_with_choices(set_a_docs)
        questions_set_b = get_questions_with_choices(set_b_docs)

        set_a_answer_key = build_answer_key(set_a_docs)
        set_b_answer_key = build_answer_key(set_b_docs)

        set_a_id = f"{board_exam_name}_{uuid.uuid4().hex[:6]}"
        set_b_id = f"{board_exam_name}_{uuid.uuid4().hex[:6]}"

        # -------------------------
        # SAVE VIA SERVICE
        # -------------------------
        TestService.create_test(set_a_id, {
            "set_id": set_a_id,
            "board_exam": board_exam_name,
            "subject": subject_name,
            "questions": questions_set_a,
            "created_at": firestore.SERVER_TIMESTAMP
        })

        TestService.create_test(set_b_id, {
            "set_id": set_b_id,
            "board_exam": board_exam_name,
            "subject": subject_name,
            "questions": questions_set_b,
            "created_at": firestore.SERVER_TIMESTAMP
        })

        TestService.create_answer_key(set_a_id, {
            "set_id": set_a_id,
            "answer_key": set_a_answer_key
        })

        TestService.create_answer_key(set_b_id, {
            "set_id": set_b_id,
            "answer_key": set_b_answer_key
        })

        # -------------------------
        # PDF RENDER
        # -------------------------
        context_a = {
            "board_exam": board_exam_name,
            "subject": subject_name,
            "questions": questions_set_a,
            "set_name": "Set A",
            "set_id": set_a_id,
            "answer_key": set_a_answer_key,
            "static_path": settings.BASE_DIR / "static",
        }

        context_b = {
            "board_exam": board_exam_name,
            "subject": subject_name,
            "questions": questions_set_b,
            "set_name": "Set B",
            "set_id": set_b_id,
            "answer_key": set_b_answer_key,
        }

        html_a = render_to_string("pdf_template.html", context_a, request=request)
        html_b = render_to_string("pdf_template.html", context_b, request=request)

        pdf_a = HTML(string=html_a).write_pdf()
        pdf_b = HTML(string=html_b).write_pdf()

        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
            zip_file.writestr(f"{set_a_id}.pdf", pdf_a)
            zip_file.writestr(f"{set_b_id}.pdf", pdf_b)

        response = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = "attachment; filename=tests.zip"
        return response

    except Exception as e:
        return HttpResponse(str(e), status=500)

def safe_created_at(x):
    ts = x.get("created_at")

    if ts is None:
        return datetime.min

    # Firestore Timestamp → datetime
    if hasattr(ts, "to_datetime"):
        return ts.to_datetime()

    return ts

def download_existing_test_pdf(request):

    set_id = request.GET.get("set_id")

    if not set_id:
        return HttpResponse("No test selected", status=400)

    try:
        test = TestService.get_test(set_id)
        answer_doc = TestService.get_answer_key(set_id)

        if not test:
            return HttpResponse("Test not found", status=404)

        answer_key = answer_doc.get("answer_key", {}) if answer_doc else {}

        context = {
            "board_exam": test.get("board_exam"),
            "subject": test.get("subject"),
            "questions": test.get("questions"),
            "set_name": set_id,
            "set_id": set_id,
            "answer_key": answer_key,
        }

        html = render_to_string("pdf_template.html", context)
        pdf = HTML(string=html).write_pdf()

        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{set_id}.pdf"'
        return response

    except Exception as e:
        return HttpResponse(str(e), status=500)




####################### FOR UPLOADING MOODLE XML FILE (QUESTIONS) TO THE QUESTION BANK  ##############################


def strip_tags(html):
    return re.sub('<[^<]+?>', '', html)


def save_question(
    question_text,
    choices,
    image_file,
    level,
    source,
    subject_name,
    topic_name,
    board_exam_list=None
):

    logger.info("🚀 save_question() CALLED")
    logger.info(f"📝 Question: {question_text[:80]}")

    # -------------------------
    # 1. IMAGE UPLOAD DEBUG
    # -------------------------
    image_path = None

    if image_file:
        try:
            image_path = upload_to_firebase(image_file, topic_name)
            logger.info(f"📦 Image stored path: {image_path}")
        except Exception:
            logger.error("❌ Image upload failed", exc_info=True)

    formatted_choices = normalize_choices(choices)
    # -------------------------
    # 3. FIRESTORE SAVE DEBUG
    # -------------------------
    try:
        logger.info("🔥 Sending to Firestore...")

        question_id = QuestionService.create_question(
            question_text=question_text,
            choices=formatted_choices,
            image_url=image_path,
            level=level,
            source=source,
            subject_names=subject_name,
            topic_name=topic_name,
            board_exam_list=board_exam_list
        )

        logger.info(f"✅ FIRESTORE SAVED: {question_id}")
        return question_id

    except Exception as e:
        logger.error("❌ Firestore save failed", exc_info=True)
        raise

def extract_and_save_questions(xml_file, subject):

    tree = ET.parse(xml_file)
    root = tree.getroot()

    logger.info("📥 XML extraction started")

    for question in root.findall('.//question'):

        question_text_element = question.find('questiontext')

        if question_text_element is not None:
            question_text, image_file = extract_question_text_and_image(
                question_text_element,
                subject
            )
        else:
            question_text = ""
            image_file = None

        answers = question.findall('answer')
        if len(answers) < 2:
            continue

        correct_answer = None
        choices = {}
        image_path = None

        for i, answer in enumerate(answers):
            text = strip_tags(answer.find('text').text.strip())
            fraction = int(answer.get('fraction'))

            letter = chr(65 + i)

            if fraction == 100:
                correct_answer = letter

            choices[letter] = (text, fraction == 100)

        if image_file:
            image_path = save_image_locally(image_file)
        
        logger.info("📄 Processing question node")
        logger.debug(f"Choices raw: {choices}")
        logger.info("💾 Calling save_question from XML")

        save_question(
            question_text=question_text,
            choices=choices,
            image_file=image_file,
            level="E",
            source="xml_import",
            subject_name=subject,
            topic_name="",
            board_exam_list=None
        )

def extract_question_text_and_image(question_text_element, subject):

    question_text = ""
    image_file = None

    text_element = question_text_element.find('./text')

    if text_element is not None:
        text_content = text_element.text or ""

        match = re.search(r'<p>(.*?)<img', text_content)
        if match:
            question_text = match.group(1).strip()
        else:
            match = re.search(r'<p>(.*?)</p>', text_content)
            if match:
                question_text = match.group(1).strip()

    file_elements = question_text_element.findall('./file')

    for file_element in file_elements:
        image_name = file_element.get('name')

        if image_name.endswith(('.png', '.jpg')):

            file_data = base64.b64decode(file_element.text.strip())

            image_file = InMemoryUploadedFile(
                ContentFile(file_data),
                None,
                image_name,
                'image/jpeg',
                len(file_data),
                None
            )

    return question_text, image_file

def save_image_locally(image_file):

    media_dir = os.path.join("media", "question_images")
    os.makedirs(media_dir, exist_ok=True)

    _, filename = os.path.split(image_file.name)
    path = os.path.join(media_dir, filename)

    with open(path, "wb") as f:
        f.write(image_file.read())

    return path


def parse_txt(text, image_files, subject_name, topic_name):

    logger.info("📥 TXT parsing started")

    lines = (text or "").splitlines()
    i = 0

    source = "google.com"

    while i < len(lines):

        line = lines[i].strip()

        if not line:
            i += 1
            continue

        # -------------------------
        # GLOBAL SOURCE
        # -------------------------
        if line.startswith("Source:"):
            source = line.replace("Source:", "").strip()
            i += 1
            continue

        # -------------------------
        # QUESTION START
        # -------------------------
        if line.startswith("<Q>"):

            question_text = line.replace("<Q>", "").strip()
            i += 1

            choices = {}
            image_name = None
            difficulty = "E"
            board_exam_list = []

            while i < len(lines) and not lines[i].startswith("<Q>"):

                l = lines[i].strip()

                if not l:
                    i += 1
                    continue

                # -------------------------
                # IMAGE
                # -------------------------
                if l.startswith("Img:"):
                    image_name = l.replace("Img:", "").strip()
                    logger.info(f"LOOKING FOR IMAGE: {image_name}")
        
                
                # -------------------------
                # CORRECT ANSWER
                # -------------------------
                elif l.startswith(">>>"):
                    clean = l.replace(">>>", "").strip()

                    # expected format: B. Blue
                    if "." in clean:
                        letter, text = clean.split(".", 1)
                        letter = letter.strip().upper()
                        text = text.strip()
                    else:
                        letter = clean[0].upper()
                        text = clean[1:].strip()

                    choices[letter] = {
                        "text": text,
                        "is_correct": True
                    }

                # -------------------------
                # NORMAL CHOICE
                # -------------------------
                elif len(l) > 1 and l[1] == ".":
                    letter = l[0].upper()
                    text = l[2:].strip()
                    choices[letter] = {"text": text, "is_correct": False}

                # -------------------------
                # DIFFICULTY
                # -------------------------
                elif l.upper() in ["VE", "E", "M", "D", "VD"]:
                    difficulty = l.upper()

                # -------------------------
                # BOARD EXAMS (AUTO DETECT)
                # -------------------------
                elif "," in l:
                    parts = [x.strip() for x in l.split(",") if x.strip()]
                    if all(len(p) <= 5 for p in parts):  # simple safety check
                        board_exam_list = parts

                i += 1

            # -------------------------
            # IMAGE (optional)
            # -------------------------
            image_file = image_files.get(image_name.lower()) if image_name else None

            logger.info(f"💾 Saving TXT question: {question_text[:50]}")
            # -------------------------
            # SAVE
            # -------------------------
            logger.info("➡️ Calling save_question from TXT")

            try:
                save_question(
                    question_text=question_text,
                    choices=choices,
                    image_file=image_file,
                    level=difficulty,
                    source=source,
                    subject_name=subject_name,
                    topic_name=topic_name,
                    board_exam_list=board_exam_list
                )
                print("✅ SAVED")
            except Exception as e:
                print("❌ ERROR SAVING QUESTION:", e)

        else:
            i += 1


def parse_xlsx(df, image_map=None, subject=None, topic=None):

    logger.info("📊 XLSX parsing started")

    # -------------------------
    # Normalize column headers
    # -------------------------
    normalized_cols = {
        "".join(str(c).lower().replace("\xa0", "").split()): c
        for c in df.columns
    }

    def get_col(name):
        key = "".join(name.lower().replace("\xa0", "").split())
        return normalized_cols.get(key)

    q_col = get_col("question")
    a_col = get_col("choicea")
    b_col = get_col("choiceb")
    c_col = get_col("choicec")
    d_col = get_col("choiced")
    e_col = get_col("choicee")
    ans_col = get_col("correctanswer")
    lvl_col = get_col("difficulty")
    img_col = get_col("image")
    src_col = get_col("source")
    be_col = get_col("boardexam")

    # -------------------------
    # ROW LOOP
    # -------------------------
    for _, row in df.iterrows():

        question_text = str(row.get(q_col, "")).strip() if q_col else ""
        source = str(row.get(src_col, "")).strip() if src_col else "unknown"

        # ✅ SAFE LEVEL HANDLING (NO CRASH EVER)
        level = "E"
        if lvl_col:
            level = str(row.get(lvl_col, "E")).strip().upper()

        # -------------------------
        # IMAGE (SAFE + CASE INSENSITIVE)
        # -------------------------
        image_name = ""
        image_file = None

        if img_col:
            image_name = str(row.get(img_col, "")).strip().lower()
            image_file = image_map.get(image_name)

        # -------------------------
        # CHOICES
        # -------------------------
        choices_raw = {
            "A": str(row.get(a_col, "")).strip() if a_col else "",
            "B": str(row.get(b_col, "")).strip() if b_col else "",
            "C": str(row.get(c_col, "")).strip() if c_col else "",
            "D": str(row.get(d_col, "")).strip() if d_col else "",
            "E": str(row.get(e_col, "")).strip() if e_col else "",
        }

        correct = str(row.get(ans_col, "")).strip().upper() if ans_col else ""

        choices = [
            {"text": v, "is_correct": (k == correct)}
            for k, v in choices_raw.items()
            if v
        ]

        # -------------------------
        # BOARD EXAMS
        # -------------------------
        board_exam_list = []
        if be_col and pd.notna(row.get(be_col)):
            board_exam_list = [p.strip() for p in str(row[be_col]).split(",")]

        # -------------------------
        # DEBUG LOGS
        # -------------------------
        logger.info(f"📝 Question: {question_text[:60]}")
        logger.info(f"📊 Level: {level}")
        logger.info(f"🖼️ Image: {image_name}")

        # -------------------------
        # SAVE
        # -------------------------
        logger.info("➡️ Calling save_question from XLSX")

        try:
            save_question(
                question_text=question_text,
                choices=choices,
                image_file=image_file,
                level=level,
                source=source,
                subject_name=subject,
                topic_name=topic,
                board_exam_list=board_exam_list
            )

        except Exception as e:
            logger.error("❌ XLSX SAVE ERROR", exc_info=True)

def normalize_choices(raw_choices):
    """
    ALWAYS returns Firestore-safe format:
    [
        {"letter": "A", "text": "...", "is_correct": False}
    ]
    """

    formatted = []

    if isinstance(raw_choices, dict):
        for i, (letter, value) in enumerate(raw_choices.items()):

            if isinstance(value, tuple):
                text, is_correct = value
            elif isinstance(value, dict):
                text = value.get("text", "")
                is_correct = value.get("is_correct", False)
            else:
                continue

            formatted.append({
                "letter": letter,
                "text": text,
                "is_correct": bool(is_correct)
            })

    elif isinstance(raw_choices, list):
        for i, c in enumerate(raw_choices):
            formatted.append({
                "letter": chr(65 + i),
                "text": c.get("text", ""),
                "is_correct": c.get("is_correct", False)
            })

    return formatted

def extract_pdf_text(file):
    text = ""
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            text += page.extract_text() or ""
    return text

def extract_text_from_docx(file):
    doc = docx.Document(file)
    return "\n".join([p.text for p in doc.paragraphs])

def extract_text_from_txt(file):
    return file.read().decode("utf-8")

def upload_to_firebase(image_file, topic_name=None):

    if not image_file:
        return None

    bucket = storage.bucket()

    safe_topic = (topic_name or "unknown").strip().replace(" ", "_")

    blob = bucket.blob(
        f"questions/{safe_topic}/{uuid.uuid4().hex}_{image_file.name}"
    )

    blob.upload_from_file(image_file)

    # ✅ IMPORTANT: return PATH ONLY (same as Add_Question view)
    return blob.name

def upload_file(request):

    if request.method == "POST":

        # -------------------------
        # 1. FIRST: GET ALL DATA SAFELY
        # -------------------------
        uploaded_items = request.FILES.getlist("folder_upload")
        subject = request.POST.get("subject", "")
        topic = request.POST.get("topic", "")

        # -------------------------
        # 2. NOW SAFE LOGGING
        # -------------------------
        logger.info("📤 upload_file triggered")
        logger.info(f"Subject: {subject}, Topic: {topic}")
        logger.info(f"Files received: {len(uploaded_items)}")

        main_file = None
        image_map = {}

        # -------------------------
        # 3. PROCESS FILES
        # -------------------------
        for f in uploaded_items:
            ext = f.name.lower().split(".")[-1]

            logger.info(f"📦 Detected file: {f.name} ({ext})")

            if ext in ["pdf", "docx", "txt", "xlsx"]:
                main_file = f
            elif ext in ["jpg", "jpeg", "png"]:
                image_map[os.path.basename(f.name).lower()] = f

        if not main_file:
            logger.error("❌ No main file found")
            return HttpResponse("No main file found")

        # -------------------------
        # 4. PARSE FILE
        # -------------------------
        try:
            ext = main_file.name.lower().split(".")[-1]

            logger.info(f"📄 Processing main file: {main_file.name}")

            if ext == "pdf":
                text = extract_pdf_text(main_file)
                parse_txt(text, image_map, subject, topic)

            elif ext == "docx":
                text = extract_text_from_docx(main_file)
                parse_txt(text, image_map, subject, topic)

            elif ext == "txt":
                text = extract_text_from_txt(main_file)
                parse_txt(text, image_map, subject, topic)

            elif ext == "xlsx":
                df = pd.read_excel(main_file)
                parse_xlsx(df, image_map=image_map, subject=subject, topic=topic)

            else:
                logger.error(f"❌ Invalid file type: {ext}")
                return HttpResponse("Invalid file type")

        except Exception as e:
            logger.error("❌ Upload processing failed", exc_info=True)
            return HttpResponse(f"Error: {str(e)}")

        logger.info("✅ Upload process completed")
        return redirect("question_bank")

    # -------------------------
    # GET REQUEST
    # -------------------------
    return render(request, "upload_file.html", {
        "BOARD_EXAMS": list(BOARD_EXAM_TOPICS.keys()),
        "BOARD_EXAM_TOPICS_JSON": json.dumps(BOARD_EXAM_TOPICS)
    })



####################### FOR UPLOADING AND CHECKING OF ANSWER SHEET (IMAGE) ##############################


def get_exam_id_suggestions(request):
    input_text = request.GET.get('input', '').lower()

    suggestions = TestService.search_answer_keys(input_text)

    return JsonResponse(suggestions, safe=False)

def get_subjects(request):
    subjects = TestService.get_all_subjects()
    return JsonResponse({"subjects": subjects})


def download_answer_page(request):
    return render(request, 'download_answer_key.html')

def download_exam_results_page(request):
    return render(request, 'download_exam_results.html')

def get_exam_dates_by_board_exam(request):
    board_exam = request.GET.get('board_exam')

    dates = TestService.get_exam_dates_by_board_exam(board_exam)

    return JsonResponse({"dates": dates})

def get_subjects_by_board_exam_and_date(request):
    board_exam = request.GET.get('board_exam')
    exam_date = request.GET.get('exam_date')

    subjects = TestService.get_subjects_by_board_exam_and_date(
        board_exam,
        exam_date
    )

    return JsonResponse({"subjects": subjects})


def download_answer_key(request):
    exam_id = request.GET.get('exam_id')

    if not exam_id:
        return JsonResponse({'error': 'Exam ID is required'})

    data = TestService.get_answer_key(exam_id)

    if not data:
        return JsonResponse({'error': 'Answer key not found'})

    file_name = f"answer_key_{exam_id}.txt"

    answer_key_str = (
        f"Board Exam/Course: {data.get('board_exam')}\n"
        f"Subject: {data.get('subject')}\n"
        f"Test Key: {exam_id}\n"
        f"{'-'*40}\n"
        + '\n'.join([f"{k}: {v}" for k, v in data.get("answer_key", {}).items()])
    )

    response = HttpResponse(answer_key_str, content_type='text/plain')
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    return response

# Get distinct board exams
def get_board_exams(request):

    exams = TestService.get_all_board_exams()

    return JsonResponse({
        "board_exams": exams
    })


# Get distinct subjects by board exam
def get_subjects_by_board_exam(request):

    board_exam = request.GET.get('board_exam')

    if not board_exam:
        return JsonResponse({"subjects": []})

    questions = QuestionService.get_all()

    subjects = set()

    for q in questions:

        board_exams = q.get("board_exams", [])

        if board_exam not in board_exams:
            continue

        for subj in q.get("subjects", []):
            if subj:
                subjects.add(subj)

    return JsonResponse({"subjects": list(subjects)})


def download_exam_results(request):
    subject = request.GET.get('subject')
    exam_date = request.GET.get('exam_date')
    board_exam = request.GET.get('board_exam')  # 

    if not subject or not exam_date or not board_exam:
        return JsonResponse({'error': 'Subject, board exam, and exam date are required'})

    try:
        month, year = exam_date.split('-')
        month_num = datetime.strptime(month, "%B").month
        year = int(year)
    except ValueError:
        return JsonResponse({'error': 'Invalid exam date format'})

    tests = TestService.get_by_subject_board_and_date(
        subject,
        board_exam,
        month_num,
        year
    )

    exam_ids = [t["id"] for t in tests]

    if not exam_ids:
        return JsonResponse({'error': 'No exams found for this subject/date'})

    # 🔥 STEP 2: Get results using service
    results = []
    for exam_id in exam_ids:
        exam_results = ResultService.get_by_exam(exam_id)
        results.extend(exam_results)

    if not results:
        return JsonResponse({'message': 'No results yet for this subject/date.'})

    # ================== EXCEL ==================
    wb = Workbook()
    ws = wb.active
    ws.title = f"{subject} Results"

    ws["A1"] = "Board Exam:"
    ws["B1"] = board_exam
    ws["A2"] = "Subject:"
    ws["B2"] = subject
    ws["A3"] = "Exam Date:"
    ws["B3"] = exam_date

    ws.append([])
    ws.append(["Student Name", "Score", "Exam Set"])

    # Styles
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for cell in ws[5]:
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    # 🔥 STEP 3: Write results
    for data in results:
        ws.append([
            data.get("student_name", ""),
            data.get("score", 0),
            data.get("exam_id", "")
        ])

    # Styling rows
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Auto width
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = f"Results_{subject}_{exam_date}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response



def image_to_mask(image):
    # Convert image to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # Apply adaptive thresholding
    _, mask = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    
    # Apply morphological operations to clean up the mask
    kernel_size = (2, 2)
    kernel = np.ones(kernel_size, np.uint8)
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
    
    return mask


def upload_answer(request):
    if request.method == 'POST' and request.FILES.get('image'):

        uploaded_image = request.FILES['image']
        exam_id = request.POST.get('exam_id')

        # --- LOAD MODELS ---
        net_original, classes_original = get_original_model()
        net_cropped, classes_cropped = get_cropped_model()

        # =========================
        # ANSWER KEY (SERVICE)
        # =========================
        answer_key_data = TestService.get_answer_key(exam_id)

        if not answer_key_data:
            return JsonResponse({"error": "Answer key not found"})

        subject = answer_key_data.get("subject")

        correct_answers = {
            str(k): v['letter']
            for k, v in answer_key_data.get("answer_key", {}).items()
        }

        # =========================
        # STUDENT (SERVICE)
        # =========================
        uid = request.session.get("uid")

        if not uid:
            return JsonResponse({"error": "Not authenticated"})

        student = UserService.get_user(uid)

        if not student:
            return JsonResponse({"error": "Student not found"})

        student_id = student.get("student_id")
        course = student.get("course")
        student_name = f"{student.get('last_name', '')} {student.get('first_name', '')}"

        # --- IMAGE DECODE ---
        nparr = np.frombuffer(uploaded_image.read(), np.uint8)
        image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

        mask_image = image_to_mask(image)

        if mask_image.ndim == 2:
            mask_image = cv2.cvtColor(mask_image, cv2.COLOR_GRAY2BGR)

        reference_point = (0, 0)

        # =========================
        # DETECTION PIPELINE
        # =========================
        boxes_original, _, class_ids_original = detect_objects(
            mask_image, net_original, classes_original
        )

        all_answers = []

        for i, box in enumerate(boxes_original):
            if classes_original[class_ids_original[i]] != 'answer':
                continue

            x, y, w, h = box
            cropped_object = mask_image[y:y+h, x:x+w]

            boxes_cropped, _, class_ids_cropped = detect_objects(
                cropped_object, net_cropped, classes_cropped
            )

            object_dict = sort_objects_by_distance(
                boxes_cropped,
                class_ids_cropped,
                classes_cropped,
                reference_point
            )

            grouped_boxes = group_and_sequence(
                object_dict.values(),
                object_dict.keys()
            )

            seq_num_class_dict = {}
            for seq_num, idx in grouped_boxes.items():
                seq_num_class_dict[seq_num] = classes_cropped[
                    class_ids_cropped[idx - 1]
                ]

            for k in sorted(seq_num_class_dict):
                all_answers.append(seq_num_class_dict[k])

        # =========================
        # SCORING
        # =========================
        score = sum(
            1 for i, a in enumerate(all_answers, start=1)
            if correct_answers.get(str(i)) == a
        )

        total_items = len(correct_answers)

        # =========================
        # DUPLICATE CHECK (SERVICE)
        # =========================
        existing = ResultService.get_by_exam(exam_id)

        if any(r for r in existing if r.get("uid") == str(uid)):
            return JsonResponse({'warning': 'Answer already uploaded for this exam.'})

        # =========================
        # SAVE RESULT (SERVICE)
        # =========================
        ResultService.create({
            "uid": str(uid),
            "student_id": student_id,
            "course": course,
            "student_name": student_name,
            "subject": subject,
            "exam_id": exam_id,
            "answer": all_answers,
            "correct_answer": list(correct_answers.values()),
            "score": score,
            "total_items": total_items,
            "is_submitted": True,
            "timestamp": firestore.SERVER_TIMESTAMP
        })

        # cleanup
        del image, mask_image, cropped_object
        gc.collect()

        return JsonResponse({
            "score": score,
            "detected": len(all_answers),
            "total_items": total_items
        })

    return render(request, 'upload_answer.html')


def answer_sheet_view(request):
    if request.method == 'POST':
        form = AnswerSheetForm(request.POST)
        if form.is_valid():
            # TODO: optionally save to Firestore later
            pass
    else:
        form = AnswerSheetForm()

    return render(request, 'answer_sheet.html', {'form': form})

# Helper: extract choice lists from questions
def extract_choices_by_letter(questions):
    letters = ['A', 'B', 'C', 'D', 'E']
    choice_map = {letter: [] for letter in letters}
    for q in questions:
        q_choices = q.get('choices', [])
        for i, letter in enumerate(letters):
            if i < len(q_choices):
                choice_map[letter].append(q_choices[i]['text'])
            else:
                choice_map[letter].append(None)  # pad empty choices
    return choice_map

def online_answer_test(request):
    if request.method != "POST":
        return render(request, "answer_test_form.html")

    now = timezone.now()

    subject = request.POST.get("subject")
    board_exam = request.POST.get("board_exam")
    exam_date_str = timezone.now().strftime("%b%Y")

    set_a_id = f"{board_exam}_{exam_date_str}_{uuid.uuid4().hex[:4]}"
    set_b_id = f"{board_exam}_{exam_date_str}_{uuid.uuid4().hex[:4]}"

    set_a_question_ids = request.POST.getlist("set_a_question_ids[]")
    set_b_question_ids = request.POST.getlist("set_b_question_ids[]")

    # -------------------------
    # FETCH VIA SERVICE
    # -------------------------
    set_a_docs = [
        QuestionService.get(qid)
        for qid in set_a_question_ids
    ]
    set_b_docs = [
        QuestionService.get(qid)
        for qid in set_b_question_ids
    ]

    set_a_docs = [q for q in set_a_docs if q]
    set_b_docs = [q for q in set_b_docs if q]

    # -------------------------
    #   IMAGE URL
    # -------------------------
    for q in set_a_docs:
        q["image"] = firebase_image_url(q.get("image"))

    for q in set_b_docs:
        q["image"] = firebase_image_url(q.get("image"))

    # -------------------------
    # FORMAT FOR TEMPLATE
    # -------------------------
    questions_set_a = get_questions_with_choices(set_a_docs)
    questions_set_b = get_questions_with_choices(set_b_docs)

    set_a_answer_key = build_answer_key(set_a_docs)
    set_b_answer_key = build_answer_key(set_b_docs)

    set_a_choice_map = extract_choices_by_letter(questions_set_a)
    set_b_choice_map = extract_choices_by_letter(questions_set_b)

    # -----------------------------
    # SAVE TESTS VIA SERVICE
    # -----------------------------
    TestService.create_test(set_a_id, {
        "set_id": set_a_id,
        "exam_date": now,
        "board_exam": board_exam,
        "subject": subject,
        "questions": questions_set_a,
        "choiceA": set_a_choice_map.get("A", []),
        "choiceB": set_a_choice_map.get("B", []),
        "choiceC": set_a_choice_map.get("C", []),
        "choiceD": set_a_choice_map.get("D", []),
        "choiceE": set_a_choice_map.get("E", [])
    })

    TestService.create_test(set_b_id, {
        "set_id": set_b_id,
        "exam_date": now,
        "board_exam": board_exam,
        "subject": subject,
        "questions": questions_set_b,
        "choiceA": set_b_choice_map.get("A", []),
        "choiceB": set_b_choice_map.get("B", []),
        "choiceC": set_b_choice_map.get("C", []),
        "choiceD": set_b_choice_map.get("D", []),
        "choiceE": set_b_choice_map.get("E", [])
    })

    TestService.create_answer_key(set_a_id, {
        "set_id": set_a_id,
        "answer_key": set_a_answer_key
    })

    TestService.create_answer_key(set_b_id, {
        "set_id": set_b_id,
        "answer_key": set_b_answer_key
    })

    # -------------------------
    # RESPONSE
    # -------------------------
    return render(request, "answer_test.html", {
        "subject": subject,
        "board_exam": board_exam,
        "set_a_questions_choices": questions_set_a,
        "set_b_questions_choices": questions_set_b,
        "set_a_id": set_a_id,
        "set_b_id": set_b_id,
    })



def answer_test_preview(request, subject, board_exam, set_a_id, set_b_id):

    # -----------------------
    # GET TEST KEYS (SERVICE)
    # -----------------------
    test_a = TestService.get_test(set_a_id)
    test_b = TestService.get_test(set_b_id)

    answer_a = TestService.get_answer_key(set_a_id)
    answer_b = TestService.get_answer_key(set_b_id)

    # -----------------------
    #   IMAGE URLS
    # -----------------------
    for q in test_a.get("questions", []):
        q["image"] = firebase_image_url(q.get("image"))

    for q in test_b.get("questions", []):
        q["image"] = firebase_image_url(q.get("image"))

    return render(request, 'answer_test.html', {
        'subject': subject,
        'board_exam': board_exam,
        'set_a_questions_choices': test_a.get("questions", []),
        'set_b_questions_choices': test_b.get("questions", []),
        'set_a_id': set_a_id,
        'set_b_id': set_b_id,
        'answer_a': answer_a,
        'answer_b': answer_b,
    })


####################### FOR ANSWERING ONLINE ##############################

import random
import json


def answer_online_exam(request):

    tests = TestService.get_all_tests()
    data = {}

    for tk in tests:

        board = (tk.get("board_exam") or "").strip().upper()
        subject = (tk.get("subject") or "").strip()
        exam_date_raw = tk.get("exam_date")
        set_id = tk.get("id")

        if not board or not subject or not set_id:
            continue

        try:
            if hasattr(exam_date_raw, "to_datetime"):
                exam_date = exam_date_raw.to_datetime()
            elif isinstance(exam_date_raw, datetime):
                exam_date = exam_date_raw
            elif isinstance(exam_date_raw, str):
                exam_date = datetime.fromisoformat(exam_date_raw)
            else:
                continue
        except:
            continue

        month_key = f"{exam_date.year}-{exam_date.month:02d}"

        data.setdefault(board, {})
        data[board].setdefault(month_key, {})
        data[board][month_key].setdefault(subject, [])

        data[board][month_key][subject].append({
            "set_id": set_id
        })

    print("\nFINAL DATA:")
    print(json.dumps(data, indent=2))

    print("\n=== DEBUG STRUCTURE ===")
    for b, months in data.items():
        print("BOARD:", b)
        for m, subjects in months.items():
            print("  MONTH:", m)
            print("  SUBJECTS:", list(subjects.keys()))

    return render(request, "answer_online_exam.html", {
        "board_exams": list(data.keys()),
        "exam_data_json": data,   # ✅ IMPORTANT  
    })


from django.shortcuts import render, redirect
from django.utils import timezone
from django.contrib import messages
from django.utils.dateparse import parse_datetime
import string, random, uuid

from .services.test_service import TestService
from .services.result_service import ResultService


def exam_form(request, set_id):

    # ===============================
    # GET TEST (SERVICE)
    # ===============================
    test_key = TestService.get_test(set_id)

    if not test_key:
        return redirect("answer_online_exam")

    test_key["set_id"] = set_id

    # ===============================
    # GET STUDENT (SESSION AUTH)
    # ===============================
    user_id = request.session.get("uid")

    if not user_id:
        return redirect("login")

    student = UserService.get_user(user_id)

    if not student:
        return redirect("login")

    # ===============================
    # ACCESS CONTROL
    # ===============================
    student_course = (student.get("course") or "").strip().lower()
    exam_code = (test_key.get("board_exam") or "").strip().upper()

    if "civil engineering" in student_course:
        allowed_exam = "CE"
    elif "mechanical engineering" in student_course:
        allowed_exam = "ME"
    elif "electrical engineering" in student_course:
        allowed_exam = "EE"
    elif "electronics engineering" in student_course:
        allowed_exam = "ECE"
    else:
        allowed_exam = None

    if allowed_exam != exam_code:
        messages.error(request, "You are not allowed to access this exam.")
        return redirect("answer_online_exam")

    # ===============================
    # PREVENT DOUBLE SUBMISSION ( ED)
    # ===============================
    results = ResultService.get_by_user(user_id)

    for r in results:
        if r.get("exam_id") == set_id and r.get("is_submitted"):
            messages.error(request, "You have already submitted this exam.")
            return redirect("warning_page")

    # ===============================
    # BUILD QUESTIONS
    # ===============================
    question_choices = []
    letters = list(string.ascii_uppercase)[:5]

    questions = test_key.get("questions", [])
    choiceA = test_key.get("choiceA", [])
    choiceB = test_key.get("choiceB", [])
    choiceC = test_key.get("choiceC", [])
    choiceD = test_key.get("choiceD", [])
    choiceE = test_key.get("choiceE", [])

    for i, q in enumerate(questions):

        question_text = q.get("question_text") or q.get("question")
        image_url = q.get("image_url") or q.get("image") or ""

        choices_raw = [
            choiceA[i] if i < len(choiceA) else "",
            choiceB[i] if i < len(choiceB) else "",
            choiceC[i] if i < len(choiceC) else "",
            choiceD[i] if i < len(choiceD) else "",
            choiceE[i] if i < len(choiceE) else "",
        ]

        # ✅ shuffle first
        random.shuffle(choices_raw)

        # ✅ enforce "None/All" to last
        choices_raw = reorder_choices(choices_raw)

        # ✅ assign letters AFTER  ing order
        choices = list(zip(letters, choices_raw))

        question_choices.append((question_text, choices, image_url))

    total_items = len(question_choices)

    MAX_ITEMS = 100
    MAX_TIME_SECONDS = 4 * 60 * 60

    total_time_limit = int((total_items / MAX_ITEMS) * MAX_TIME_SECONDS)
    per_question_time_limit = total_time_limit / max(total_items, 1)

    # ===============================
    # POST (SUBMIT EXAM)
    # ===============================
    if request.method == "POST":

        if request.session.get("form_submitted"):
            messages.error(request, "You already submitted this exam.")
            return redirect("warning_page")

        submitted_answers = []

        for i in range(total_items):
            ans = request.POST.get(f"question_{i + 1}")

            if not ans:
                messages.error(request, f"Answer missing for question {i + 1}")
                return render(request, "exam_form.html", {
                    "test_key": test_key,
                    "question_choices": question_choices,
                    "total_items": total_items,
                    "total_time_limit": total_time_limit,
                    "per_question_time_limit": per_question_time_limit,
                    "start_time": timezone.now().isoformat(),
                })

            submitted_answers.append(ans)

        # ===============================
        # TIME TRACKING
        # ===============================
        elapsed_time = None
        start_time_str = request.POST.get("start_time")

        if start_time_str:
            start_time = parse_datetime(start_time_str)
            if start_time:
                diff = timezone.now() - start_time
                h, r = divmod(int(diff.total_seconds()), 3600)
                m, s = divmod(r, 60)
                elapsed_time = f"{h}hr {m}min {s}sec"

        request.session["form_submitted"] = True

        # ===============================
        # SCORING (SERVICE)
        # ===============================
        answer_data = TestService.get_answer_key(set_id)

        if not answer_data:
            return redirect("warning_page")

        answer_key = answer_data.get("answer_key", {})

        score = 0
        correct_answers = []

        for i, key in enumerate(sorted(answer_key.keys(), key=int)):
            correct = answer_key[key].get("text", "")
            correct_answers.append(correct)

            if i < len(submitted_answers) and submitted_answers[i] == correct:
                score += 1

        # ===============================
        # SAVE RESULT (SERVICE)
        # ===============================
        result_id = ResultService.create({
            "user_id": user_id,
            "student_id": student.get("student_id"),
            "course": student.get("course"),
            "student_name": student.get("first_name", "") + " " + student.get("last_name", ""),
            "subject": test_key.get("subject"),
            "exam_id": set_id,
            "answer": submitted_answers,
            "correct_answer": correct_answers,
            "score": score,
            "total_items": total_items,
            "is_submitted": True,
            "timestamp": timezone.now().isoformat(),
            "elapsed_time": elapsed_time
        })

        return redirect("result_page", result_id=result_id)

    # ===============================
    # GET
    # ===============================
    return render(request, "exam_form.html", {
        "test_key": test_key,
        "question_choices": question_choices,
        "total_items": total_items,
        "total_time_limit": total_time_limit,
        "per_question_time_limit": per_question_time_limit,
        "start_time": timezone.now().isoformat(),
    })


def result_page(request, result_id):

    result = ResultService.get_by_id(result_id)

    if not result:
        return render(request, "result_page.html", {"error": "Result not found"})

    percent = 0
    if result.get("total_items"):
        percent = round(
            (result.get("score", 0) / result.get("total_items")) * 100,
            2
        )

    return render(request, "result_page.html", {
        "result": result,
        "percent": percent
    })


def warning_page(request):
    home_student_url = reverse('home_student')  # Assuming 'home_student' is the name of the URL pattern for home_student.html
    return render(request, 'submit_warning.html', {'home_student_url': home_student_url})

def view_results(request):

    # =========================
    # GET USER FROM SESSION (CONSISTENT WITH PRACTICE_START)
    # =========================
    user_id = request.session.get("uid")

    if not user_id:
        return redirect("login")

    student = UserService.get_user(user_id)

    if not student:
        return redirect("login")

    # =========================
    # GET RESULTS
    # =========================
    results = ResultService.get_by_user(user_id)

    # sort by timestamp (latest first)
    results.sort(
        key=lambda x: x.get("timestamp", ""),
        reverse=True
    )

    return render(request, "view_results.html", {
        "results": results
    })

def question_analytics(request):

    questions = QuestionService.get_all()

    board_exam_counts = {}
    subject_counts = {}
    difficulty_counts = {}

    for data in questions:

        board_exams = data.get("board_exams", [])
        subjects = data.get("subjects", [])
        difficulty = data.get("difficulty", "Unknown Difficulty")

        if not board_exams:
            board_exams = ["No Board Exam"]

        if not subjects:
            subjects = ["Unknown Subject"]

        # -------------------
        # BOARD EXAM COUNTS
        # -------------------
        for be in board_exams:
            board_exam_counts[be] = board_exam_counts.get(be, 0) + 1

            for subj in subjects:
                key = (be, subj)
                subject_counts[key] = subject_counts.get(key, 0) + 1

        # -------------------
        # DIFFICULTY COUNTS
        # -------------------
        difficulty_counts[difficulty] = difficulty_counts.get(difficulty, 0) + 1

    # -------------------
    # FORMAT FOR TEMPLATE
    # -------------------
    course_stats = [
        {"board_exam": k, "total": v}
        for k, v in board_exam_counts.items()
    ]

    subject_distribution = [
        {
            "board_exam": be,
            "subject": subj,
            "total_questions": count
        }
        for (be, subj), count in subject_counts.items()
    ]

    context = {
        "course_stats": course_stats,
        "subject_distribution": subject_distribution,

        "course_labels": json.dumps(list(board_exam_counts.keys())),
        "total_questions": json.dumps(list(board_exam_counts.values())),

        "difficulty_labels": json.dumps(list(difficulty_counts.keys())),
        "difficulty_counts": json.dumps(list(difficulty_counts.values())),
    }

    return render(request, "question_analytics.html", context)

def test_analytics(request):

    # ✅ USE SERVICE
    results = ResultService.get_all()

    course_results_map = {}

    for r in results:
        course = r.get("course")

        if not course:
            continue

        course_results_map.setdefault(course, []).append(r)

    course_data = {}

    for course, results in course_results_map.items():

        passed_counts = 0
        failed_counts = 0
        total_score = 0

        question_stats = defaultdict(lambda: {'correct': 0, 'wrong': 0})

        for r in results:
            score = r.get("score", 0)
            total_items = r.get("total_items", 1)

            if score >= 0.6 * total_items:
                passed_counts += 1
            else:
                failed_counts += 1

            total_score += score

            answers = r.get("answer", []) or []
            correct_answers = r.get("correct_answer", []) or []

            for idx, ans in enumerate(answers):
                correct_ans = correct_answers[idx] if idx < len(correct_answers) else None

                if ans == correct_ans:
                    question_stats[idx]['correct'] += 1
                else:
                    question_stats[idx]['wrong'] += 1

        avg_score = total_score / len(results) if results else 0

        question_labels = [f'Q{i+1}' for i in range(len(question_stats))]
        correct_counts = [question_stats[i]['correct'] for i in range(len(question_stats))]
        wrong_counts = [question_stats[i]['wrong'] for i in range(len(question_stats))]

        top_students = sorted(results, key=lambda x: x.get("score", 0), reverse=True)[:10]

        course_data[course] = {
            "passed_counts": passed_counts,
            "failed_counts": failed_counts,
            "avg_score": avg_score,
            "question_labels": question_labels,
            "correct_counts": correct_counts,
            "wrong_counts": wrong_counts,
            "top_students": top_students,
            "passed_json": json.dumps({"passed": passed_counts, "failed": failed_counts}),
            "avg_json": json.dumps({"avg": avg_score}),
            "question_json": json.dumps({
                "labels": question_labels,
                "correct": correct_counts,
                "wrong": wrong_counts
            }),
        }

    return render(request, "test_analytics.html", {
        "course_data": course_data
    })

def build_practice_set(board_exam, subject, num_items):

    questions = QuestionService.get_by_board_and_subject(board_exam, subject)

    if not questions:
        return None

    random.shuffle(questions)

    selected = questions[:num_items]

    payload = []

    for q in selected:

        choices = q.get("choices", [])

        # ✅  : always push special choices to last (E)
        choices = reorder_choices(choices)

        # ensure correct key mapping (IMPORTANT  )
        correct_letter = q.get("correct_letter") or q.get("correct_answer")

        payload.append({
            "id": q["id"],
            "text": q.get("question_text"),
            "image_name": q.get("image_name") or q.get("image"),
            "choices": choices,
            "correct": correct_letter,
            "subject": subject,
            "topic": q.get("topic", "Misc"),
            "difficulty": q.get("difficulty", "Unknown"),
        })

    return payload

# ---- Practice: start ----
@require_http_methods(["GET", "POST"])
def practice_start(request):

    # =========================
    # GET USER FROM SESSION
    # =========================
    user_id = request.session.get("uid")

    if not user_id:
        return redirect("login")

    student = UserService.get_user(user_id)

    if not student:
        return redirect("login")

    student_course_full = (student.get("course") or "").strip().lower()

    course_mapping = {
        "civil engineering": "CE",
        "mechanical engineering": "ME",
        "electrical engineering": "EE",
        "electronics engineering": "ECE",
    }

    student_course_code = course_mapping.get(student_course_full)

    if not student_course_code:
        messages.error(request, "Your course is not supported for practice exams.")
        return redirect("home")

    subjects = list(BOARD_EXAM_TOPICS.get(student_course_code, {}).keys())

    # =========================
    # POST
    # =========================
    if request.method == "POST":

        subject_name = request.POST.get("subject")

        try:
            num_items = int(request.POST.get("num_items") or 5)
        except ValueError:
            num_items = 5

        # =========================
        # 🔥 CACHE BUILDER USED HERE
        # =========================
        payload = build_practice_set(
            student_course_code,
            subject_name,
            num_items
        )

        if not payload:
            messages.error(request, "No questions found for this subject!")
            return redirect("practice_start")

        session_id = str(uuid.uuid4())

        request.session[f"practice_{session_id}"] = {
            "board_exam": student_course_code,
            "subject": subject_name,
            "questions": payload,
            "total_items": len(payload),
        }

        request.session.modified = True

        return redirect("practice_take", session_id=session_id)

    return render(request, "practice_start.html", {
        "subjects": subjects,
        "student_course": student_course_code,
    })


# ---- Practice: take (render questions + timer) ----
def practice_take(request, session_id):

    sess_key = f'practice_{session_id}'
    data = request.session.get(sess_key)

    if not data:
        messages.error(request, "Practice session not found or expired.")
        return redirect('practice_start')

    letters = list(string.ascii_uppercase)

    questions_for_client = []

    for qi, q in enumerate(data["questions"], start=1):

        choices = q["choices"].copy()

        # ✅ STEP 1: shuffle choices
        random.shuffle(choices)

        # ✅ STEP 2: enforce "None/All" to be last (E)
        choices = reorder_choices(choices)

        # ✅ STEP 3: assign letters AFTER  ing order
        for idx, choice in enumerate(choices):
            choice["display_letter"] = letters[idx]

        questions_for_client.append({
            "instance_id": qi,
            "q_id": q["id"],
            "text": q["text"],
            "image_name": q.get("image_name"),
            "choices": choices
        })

    return render(request, "practice_take.html", {
        "session_id": session_id,
        "board_exam": data["board_exam"],
        "questions": questions_for_client,
        "total_items": data["total_items"],
    })


# ---- Practice: submit (grade & analytics) ----
@require_http_methods(["POST"])
def practice_submit(request, session_id):

    sess_key = f'practice_{session_id}'
    data = request.session.get(sess_key)

    if not data:
        messages.error(request, "Practice session not found or expired.")
        return redirect('practice_start')

    questions = data['questions']
    total_items = len(questions)

    results = []
    correct_count = 0
    total_time_elapsed = 0.0

    subject_tracker = {}
    topic_tracker = {}
    difficulty_tracker = {}

    # =========================
    # PROCESS ANSWERS
    # =========================
    for i, q in enumerate(questions, start=1):

        ans = request.POST.get(f'answer_{i}')
        time_spent = float(request.POST.get(f'time_{i}', '0') or 0)

        correct_key = q.get('correct')
        is_correct = (ans == correct_key)

        subject = q.get("subject", "Unknown")
        topic = q.get("topic", "Misc")
        difficulty = q.get("difficulty", "Unknown")

        # SUBJECT
        subject_tracker.setdefault(subject, {"correct": 0, "total": 0, "time": 0})
        subject_tracker[subject]["total"] += 1
        subject_tracker[subject]["time"] += time_spent
        if is_correct:
            subject_tracker[subject]["correct"] += 1

        # TOPIC
        topic_tracker.setdefault(topic, {"correct": 0, "total": 0, "time": 0})
        topic_tracker[topic]["total"] += 1
        topic_tracker[topic]["time"] += time_spent
        if is_correct:
            topic_tracker[topic]["correct"] += 1

        # DIFFICULTY
        difficulty_tracker.setdefault(difficulty, {"correct": 0, "total": 0, "time": 0})
        difficulty_tracker[difficulty]["total"] += 1
        if is_correct:
            difficulty_tracker[difficulty]["correct"] += 1

        if is_correct:
            correct_count += 1

        total_time_elapsed += time_spent

        results.append({
            "index": i,
            "q_id": q["id"],
            "text": q["text"],
            "selected": ans,
            "correct": correct_key,
            "is_correct": is_correct,
            "time_spent": time_spent,
        })

    score = correct_count
    pct = (score / total_items * 100) if total_items else 0

    user_id = request.session.get("uid")

    # =========================
    # ANALYTICS UPDATE (SERVICE)
    # =========================
    for subject, stats in subject_tracker.items():
        PracticeService.update_analytics(
            "subject_analytics",
            f"{user_id}_{subject}_{data['board_exam']}",
            {
                "correct": stats["correct"],
                "total": stats["total"],
                "time": stats["time"]
            }
        )

    for topic, stats in topic_tracker.items():
        PracticeService.update_analytics(
            "topic_analytics",
            f"{user_id}_{topic}",
            {
                "correct": stats["correct"],
                "total": stats["total"],
                "time": stats["time"]
            }
        )

    for difficulty, stats in difficulty_tracker.items():
        PracticeService.update_analytics(
            "difficulty_analytics",
            f"{user_id}_{difficulty}_{data['board_exam']}",
            {
                "correct": stats["correct"],
                "total": stats["total"],
                "time": stats["time"]
            }
        )
    # =========================
    # SAVE RESULT (SERVICE)
    # =========================
    PracticeService.save_result(session_id, {
        "session_id": session_id,
        "user_id": user_id,
        "board_exam": data["board_exam"],
        "total_items": total_items,
        "score": score,
        "percent": pct,
        "total_time": total_time_elapsed,
        "answers": results,
    })

    # =========================
    # STORE SESSION RESULT
    # =========================
    request.session[f'practice_result_{session_id}'] = {
        "score": score,
        "total_items": total_items,
        "percent": pct,
        "results": results,
        "total_time": total_time_elapsed,
        "board_exam": data["board_exam"],
    }

    request.session.modified = True

    return redirect('practice_result_page', session_id=session_id)

def practice_result_page(request, session_id):

    res = PracticeService.get_result(session_id)

    if not res:
        messages.error(request, "No practice results found for that session.")
        return redirect('practice_start')

    return render(request, 'practice_result.html', {
        'res': res
    })

from django.core.serializers.json import DjangoJSONEncoder

@firebase_login_required
def analytics_dashboard(request):

    user_id = request.session.get("uid")

    if not user_id:
        return redirect("login")

    results = ResultService.get_by_user(user_id)

    subject_data = {}
    topic_data = {}
    difficulty_data = {}

    for res in results:

        board_exam = res.get("board_exam")
        answers = res.get("answers", [])

        for ans in answers:

            q_id = ans.get("q_id")
            selected = ans.get("selected")
            correct = ans.get("correct")
            time_spent = ans.get("time_spent", 0.0)

            # =========================
            # QUESTION FETCH (SERVICE)
            # =========================
            q = QuestionService.get(q_id)
            if not q:
                continue

            subjects = q.get("subjects", [])
            topic = q.get("topic", "Misc")
            difficulty = q.get("difficulty", "Unknown")

            subject = "Unknown"
            for subj in subjects:
                if subj in BOARD_EXAM_TOPICS.get(board_exam, {}):
                    subject = subj
                    break

            # ================= SUBJECT =================
            key = (subject, board_exam)
            subject_data.setdefault(key, {
                "total_items_answered": 0,
                "total_correct": 0,
                "total_time": 0.0,
            })

            subject_data[key]["total_items_answered"] += 1
            subject_data[key]["total_correct"] += int(selected == correct)
            subject_data[key]["total_time"] += time_spent

            # ================= TOPIC =================
            key_topic = (topic, subject)
            topic_data.setdefault(key_topic, {
                "total_items_answered": 0,
                "total_correct": 0,
                "total_time": 0.0,
            })

            topic_data[key_topic]["total_items_answered"] += 1
            topic_data[key_topic]["total_correct"] += int(selected == correct)
            topic_data[key_topic]["total_time"] += time_spent

            # ================= DIFFICULTY =================
            key_diff = (difficulty, board_exam)
            difficulty_data.setdefault(key_diff, {
                "total_items_answered": 0,
                "total_correct": 0,
            })

            difficulty_data[key_diff]["total_items_answered"] += 1
            difficulty_data[key_diff]["total_correct"] += int(selected == correct)

    # =========================
    # BUILD SUBJECT LIST
    # =========================
    subject_list = []
    for (subj, be), v in subject_data.items():

        avg_time = v["total_time"] / v["total_items_answered"] if v["total_items_answered"] else 0
        acc = (v["total_correct"] / v["total_items_answered"] * 100) if v["total_items_answered"] else 0

        subject_list.append({
            "subject": subj,
            "board_exam": be,
            "total_items_answered": v["total_items_answered"],
            "total_correct": v["total_correct"],
            "average_time_per_item": round(avg_time, 2),
            "accuracy": round(acc, 2),
        })

    # =========================
    # BUILD TOPIC LIST
    # =========================
    topic_list = []
    for (topic, subj), v in topic_data.items():

        avg_time = v["total_time"] / v["total_items_answered"] if v["total_items_answered"] else 0
        acc = (v["total_correct"] / v["total_items_answered"] * 100) if v["total_items_answered"] else 0

        topic_list.append({
            "subject": subj,
            "topic": topic,
            "total_items_answered": v["total_items_answered"],
            "total_correct": v["total_correct"],
            "average_time_per_item": round(avg_time, 2),
            "accuracy": round(acc, 2),
        })

    # =========================
    # BUILD DIFFICULTY LIST
    # =========================
    difficulty_list = []
    for (diff, be), v in difficulty_data.items():

        acc = (v["total_correct"] / v["total_items_answered"] * 100) if v["total_items_answered"] else 0

        difficulty_list.append({
            "difficulty": diff,
            "board_exam": be,
            "total_items_answered": v["total_items_answered"],
            "total_correct": v["total_correct"],
            "accuracy": round(acc, 2),
        })

    return render(request, "analytics_dashboard.html", {
        "subject_analytics": json.dumps(subject_list),
        "topic_analytics": json.dumps(topic_list),
        "difficulty_analytics": json.dumps(difficulty_list),
        "ai_suggestions": "",
    })
