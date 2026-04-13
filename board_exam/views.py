from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from .models import QuestionForm, Question, BoardExam, Subject, Topic, DifficultyLevel, Question, QuestionImage, Choice, AnswerKey, TestKey, Teacher, Student, Result, PracticeResult, SubjectAnalytics, TopicAnalytics, DifficultyAnalytics
from django.views import View
from django.template.loader import render_to_string
from weasyprint import HTML
import random, io, zipfile, uuid, re
import xml.etree.ElementTree as ET
from scripts.check import detect_objects, sort_objects_by_distance, group_and_sequence
from django.http import JsonResponse
import numpy as np
import cv2, time, os, json, base64, traceback
from django.contrib.auth import logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from django.views.decorators.csrf import csrf_protect
from django.db import IntegrityError
from django.db.models import Q
from .forms import AnswerSheetForm, SignUpForm, EmailAuthenticationForm 
from itertools import zip_longest
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.utils import timezone
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.conf import settings
import PyPDF2
import fitz, gc
import pdfplumber, docx, string
from docx import Document
from PyPDF2 import PdfReader
from django.core.files import File
import pandas as pd
from .config import BOARD_EXAM_TOPICS, LEVELS
from django.db import models, transaction
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.db.models import Count, Q, Avg, F
from collections import defaultdict
from django.views.decorators.http import require_http_methods
import datetime
import openai
from datetime import datetime
# from scripts.model_loader import net_original, classes_original, net_cropped, classes_cropped
from scripts.model_loader import get_original_model, get_cropped_model



logo_path = os.path.join(settings.BASE_DIR, 'static', 'EXIM2.png')  # full path
# SET_ID_PREFIX = {
#     "Civil Engineering": "CE",
#     "Mechanical Engineering": "ME",
#     "Electronics Engineering": "ECE",
#     "Electrical Engineering": "EE",
# }

from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponseForbidden
from django.contrib import messages
from django.views.decorators.csrf import csrf_protect
from django.contrib.auth.decorators import login_required

from firebase_admin import firestore, auth

db = firestore.client()

# =========================
# FIRESTORE TEST
# =========================
def test_firestore(request):
    doc_ref = db.collection("test").add({
        "message": "Cloud Run can access Firestore",
        "status": "success",
        "timestamp": firestore.SERVER_TIMESTAMP
    })

    return JsonResponse({
        "ok": True,
        "doc_id": doc_ref[1].id
    })


# =========================
# SIGNUP (FIREBASE AUTH)
# =========================
@csrf_protect
def signup(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)

        if not form.is_valid():
            return render(request, 'signup.html', {'form': form})

        try:
            role = form.cleaned_data['role']
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']

            # 1. Create Firebase Auth user
            user_record = auth.create_user(
                email=email,
                password=password
            )

            uid = user_record.uid

            # 2. Create Firestore user profile
            db.collection("users").document(uid).set({
                "email": email,
                "role": role,
                "first_name": form.cleaned_data.get('first_name', ''),
                "last_name": form.cleaned_data.get('last_name', ''),
                "middle_name": form.cleaned_data.get('middle_name', ''),
                "student_id": form.cleaned_data.get('student_id', ''),
                "course": form.cleaned_data.get('course', ''),
                "birthdate": str(form.cleaned_data.get('birthdate', '')),
                "created_at": firestore.SERVER_TIMESTAMP
            })

            messages.success(request, "Account created successfully!")
            return redirect('login')

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            return render(request, 'signup.html', {'form': form})

    return render(request, 'signup.html', {'form': SignUpForm()})


# =========================
# LOGIN (FIREBASE TOKEN)
# =========================
def login_view(request):
    if request.method == 'POST':
        id_token = request.POST.get('id_token')

        try:
            decoded = auth.verify_id_token(id_token)
            uid = decoded['uid']

            user_doc = db.collection("users").document(uid).get()

            if not user_doc.exists:
                return JsonResponse({"error": "User not found"}, status=404)

            user_data = user_doc.to_dict()
            role = user_data.get("role")

            # store session
            request.session["uid"] = uid
            request.session["role"] = role

            if role == "teacher":
                return redirect("home")
            else:
                return redirect("home_student")

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return render(request, 'login.html')


# =========================
# LOGOUT
# =========================
def logout_view(request):
    request.session.flush()
    return redirect('login')


# =========================
# ROLE HELPERS
# =========================
def get_user_role(request):
    uid = request.session.get("uid")
    if not uid:
        return None

    user_doc = db.collection("users").document(uid).get()

    if user_doc.exists:
        return user_doc.to_dict().get("role")

    return None


# =========================
# DASHBOARD ROUTING
# =========================
@login_required
def main_dashboard(request):
    role = request.session.get("role")

    if role == "teacher":
        return redirect("home")
    elif role == "student":
        return redirect("home_student")
    else:
        return redirect("login")


# =========================
# TEACHER DASHBOARD
# =========================
@login_required
def home(request):
    role = request.session.get("role")

    if role != "teacher":
        return HttpResponseForbidden("You cannot access this page")

    return render(request, 'home.html')


# =========================
# STUDENT DASHBOARD
# =========================
@login_required
def home_student(request):
    role = request.session.get("role")

    if role != "student":
        return HttpResponseForbidden("You cannot access this page")

    return render(request, 'home_student.html')


# =========================
# ROOT REDIRECT
# =========================
def root_redirect(request):
    if not request.session.get("uid"):
        return redirect("login")

    role = request.session.get("role")

    if role == "teacher":
        return redirect("home")
    elif role == "student":
        return redirect("home_student")
    else:
        return redirect("login")

from firebase_admin import firestore
from google.cloud import storage
import random, uuid

db = firestore.client()

storage_client = storage.Client()
BUCKET_NAME = "exim-media-concrete-potion-477505-p2"
bucket = storage_client.bucket(BUCKET_NAME)

def serve_image(request, image_name):
    blob = bucket.blob(image_name)

    if not blob.exists():
        return HttpResponse("Image not found", status=404)

    image_bytes = blob.download_as_bytes()
    content_type = blob.content_type or "application/octet-stream"

    return HttpResponse(image_bytes, content_type=content_type)

def question_bank(request):
    questions_ref = db.collection("questions").stream()

    questions = []

    for doc in questions_ref:
        q = doc.to_dict()
        q["id"] = doc.id

        # format choices
        letters = ["A", "B", "C", "D", "E"]
        choices = q.get("choices", [])

        q["lettered_choices"] = [
            {"letter": letters[i], "text": c.get("text")}
            for i, c in enumerate(choices[:5])
        ]

        q["correct_answer_text"] = next(
            (c["text"] for c in choices if c.get("is_correct")),
            "-"
        )

        q["image_names"] = q.get("images", [])

        questions.append(q)

    return render(request, "question_bank.html", {"questions": questions})

from django.views import View
from django.shortcuts import redirect

class Add_Question(View):
    def get(self, request):
        context = {
            'BOARD_EXAMS': list(BOARD_EXAM_TOPICS.keys()),
            'LEVELS_JSON': json.dumps(LEVELS),
        }
        return render(request, "add_question.html", context)

    def post(self, request):

        num_questions = sum(
            1 for key in request.POST.keys()
            if key.startswith("question_text_")
        )

        if num_questions == 0:
            messages.error(request, "No questions to save!")
            return redirect("add_question")

        for i in range(1, num_questions + 1):

            board_exam_names = request.POST.getlist(f"board_exam_{i}")
            subject_names = request.POST.getlist(f"subjects_{i}[]")
            topic_name = request.POST.get(f"topic_{i}")
            level_name = request.POST.get(f"level_{i}")
            question_text = request.POST.get(f"question_text_{i}")
            source = request.POST.get(f"source_{i}", "google.com")

            # upload image
            image_file = request.FILES.get(f"image_{i}")
            image_url = None

            if image_file:
                blob = bucket.blob(f"questions/{uuid.uuid4().hex}.jpg")
                blob.upload_from_file(image_file)
                blob.make_public()
                image_url = blob.public_url

            # build choices
            choices = []
            correct_letter = request.POST.get(f"correct_answer_{i}")

            for letter in ["A", "B", "C", "D", "E"]:
                text = request.POST.get(f"choice{letter}_{i}")
                if text:
                    choices.append({
                        "text": text,
                        "is_correct": letter == correct_letter
                    })

            # FIRESTORE QUESTION
            db.collection("questions").add({
                "board_exams": board_exam_names,
                "subjects": subject_names,
                "topic": topic_name,
                "difficulty": level_name,
                "question_text": question_text,
                "source": source,
                "images": [image_url] if image_url else [],
                "choices": choices,
                "usage_count": 0,
                "created_at": firestore.SERVER_TIMESTAMP
            })

        messages.success(request, f"{num_questions} questions added successfully!")
        return redirect("question_bank")

def get_random_questions(num_questions, subject):

    all_q = db.collection("questions").where(
        "subjects", "array_contains", subject
    ).stream()

    questions = [q.to_dict() | {"id": q.id} for q in all_q]

    if num_questions > len(questions):
        raise ValueError("Not enough questions")

    return random.sample(questions, num_questions)

def generate_set_id(board_exam):
    prefix_map = {
        "civil": "CE",
        "mechanical": "ME",
        "electronics": "ECE",
        "electrical": "EE"
    }

    board_exam_lower = board_exam.lower()

    prefix = "GEN"
    for k, v in prefix_map.items():
        if k in board_exam_lower:
            prefix = v

    return f"{prefix}_{uuid.uuid4().hex[:8]}"
    
def generate_test(request):

    if request.method == "POST":

        board_exam = request.POST.get("board_exam")
        subject = request.POST.get("subject", "")
        num_questions = int(request.POST.get("num_questions", 0))

        qs = db.collection("questions")

        if board_exam:
            qs = qs.where("board_exams", "array_contains", board_exam)

        if subject:
            qs = qs.where("subjects", "array_contains", subject)

        docs = list(qs.stream())

        questions = [d.to_dict() | {"id": d.id} for d in docs]

        if len(questions) < num_questions:
            return render(request, "generate_test.html", {
                "error_message": "Not enough questions"
            })

        selected = random.sample(questions, num_questions)

        return render(request, "generated_test.html", {
            "set_a_questions": selected,
            "set_b_questions": random.sample(selected, len(selected)),
            "board_exam": board_exam,
            "subject": subject,
            "set_a_id": uuid.uuid4().hex,
            "set_b_id": uuid.uuid4().hex,
        })

    return render(request, "generate_test.html")
    
import string
import random
import uuid
from firebase_admin import firestore

db = firestore.client()

def map_letter_text(choices_lists, correct_text_dict):

    answer_key = {}
    num_choices = len(choices_lists)
    letters = list(string.ascii_uppercase[:num_choices])

    for i, correct_text in correct_text_dict.items():

        choice_map = {
            letters[idx]: choices_lists[idx][i-1]
            for idx in range(num_choices)
        }

        correct_letter = next(
            (l for l, t in choice_map.items() if t == correct_text),
            None
        )

        answer_key[str(i)] = {
            "letter": correct_letter,
            "text": correct_text
        }

    return answer_key

def get_questions_with_choices(question_docs):

    questions = []
    letters = ['A', 'B', 'C', 'D', 'E']

    for doc in question_docs:
        q = doc.to_dict()

        choices = q.get("choices", [])

        formatted_choices = []
        for i, c in enumerate(choices[:5]):
            formatted_choices.append({
                "letter": letters[i],
                "text": c.get("text")
            })

        questions.append({
            "id": doc.id,
            "question": q.get("question_text"),
            "choices": formatted_choices,
            "image_url": q.get("images", [None])[0]
        })

    return questions

def build_answer_key(question_docs):

    letters = ['A', 'B', 'C', 'D', 'E']
    answer_key = {}

    for i, doc in enumerate(question_docs, start=1):

        q = doc.to_dict()
        choices = q.get("choices", [])

        for idx, c in enumerate(choices):
            if c.get("is_correct"):
                answer_key[str(i)] = {
                    "letter": letters[idx],
                    "text": c.get("text")
                }
                break

    return answer_key


def download_test_pdf(request):

    if request.method != "POST":
        return HttpResponse("Invalid request method", status=405)

    try:
        subject_name = request.POST.get("subject")
        board_exam_name = request.POST.get("board_exam")

        set_a_ids = request.POST.getlist("set_a_question_ids[]")
        set_b_ids = request.POST.getlist("set_b_question_ids[]")

        # 🔥 FETCH FROM FIRESTORE
        set_a_docs = [
            db.collection("questions").document(qid).get()
            for qid in set_a_ids
        ]

        set_b_docs = [
            db.collection("questions").document(qid).get()
            for qid in set_b_ids
        ]

        questions_set_a = get_questions_with_choices(set_a_docs)
        questions_set_b = get_questions_with_choices(set_b_docs)

        set_a_answer_key = build_answer_key(set_a_docs)
        set_b_answer_key = build_answer_key(set_b_docs)

        set_a_id = f"{board_exam_name}_{uuid.uuid4().hex[:6]}"
        set_b_id = f"{board_exam_name}_{uuid.uuid4().hex[:6]}"

        # 💾 SAVE TO FIRESTORE (instead of SQL)
        db.collection("test_keys").document(set_a_id).set({
            "set_id": set_a_id,
            "board_exam": board_exam_name,
            "subject": subject_name,
            "questions": questions_set_a,
            "choices": {q["id"]: q["choices"] for q in questions_set_a},
            "created_at": firestore.SERVER_TIMESTAMP
        })

        db.collection("test_keys").document(set_b_id).set({
            "set_id": set_b_id,
            "board_exam": board_exam_name,
            "subject": subject_name,
            "questions": questions_set_b,
            "choices": {q["id"]: q["choices"] for q in questions_set_b},
            "created_at": firestore.SERVER_TIMESTAMP
        })

        # 🔑 SAVE ANSWER KEYS
        db.collection("answer_keys").document(set_a_id).set({
            "set_id": set_a_id,
            "answer_key": set_a_answer_key
        })

        db.collection("answer_keys").document(set_b_id).set({
            "set_id": set_b_id,
            "answer_key": set_b_answer_key
        })

        # 📄 RENDER PDF
        context_a = {
            "board_exam": board_exam_name,
            "subject": subject_name,
            "questions": questions_set_a,
            "set_name": "Set A",
            "set_id": set_a_id,
            "answer_key": set_a_answer_key,
        }

        context_b = {
            "board_exam": board_exam_name,
            "subject": subject_name,
            "questions": questions_set_b,
            "set_name": "Set B",
            "set_id": set_b_id,
            "answer_key": set_b_answer_key,
        }

        html_a = render_to_string("pdf_template.html", context_a, request=request)
        html_b = render_to_string("pdf_template.html", context_b, request=request)

        pdf_a = HTML(string=html_a).write_pdf()
        pdf_b = HTML(string=html_b).write_pdf()

        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
            zip_file.writestr(f"{set_a_id}.pdf", pdf_a)
            zip_file.writestr(f"{set_b_id}.pdf", pdf_b)

        response = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = "attachment; filename=tests.zip"
        return response

    except Exception as e:
        return HttpResponse(str(e), status=500)

def download_existing_test_pdf(request):

    set_id = request.GET.get("set_id")

    if not set_id:
        return HttpResponse("No test selected", status=400)

    try:
        test_doc = db.collection("test_keys").document(set_id).get()
        answer_doc = db.collection("answer_keys").document(set_id).get()

        if not test_doc.exists:
            return HttpResponse("Test not found", status=404)

        test = test_doc.to_dict()
        answer_key = answer_doc.to_dict().get("answer_key", {}) if answer_doc.exists else {}

        context = {
            "board_exam": test.get("board_exam"),
            "subject": test.get("subject"),
            "questions": test.get("questions"),
            "set_name": set_id,
            "set_id": set_id,
            "answer_key": answer_key,
        }

        html = render_to_string("pdf_template.html", context)
        pdf = HTML(string=html).write_pdf()

        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{set_id}.pdf"'
        return response

    except Exception as e:
        return HttpResponse(str(e), status=500)




####################### FOR UPLOADING MOODLE XML FILE (QUESTIONS) TO THE QUESTION BANK  ##############################

from firebase_admin import firestore
from google.cloud import storage
import os, base64, re, uuid
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import InMemoryUploadedFile

db = firestore.client()

storage_client = storage.Client()
BUCKET_NAME = "exim-media-concrete-potion-477505-p2"
bucket = storage_client.bucket(BUCKET_NAME)


def strip_tags(html):
    return re.sub('<[^<]+?>', '', html)


def save_question_firestore(
    question_text,
    choices,
    image_file,
    level,
    source,
    subject_name,
    topic_name,
    board_exam_list=None
):

    # upload image to Cloud Storage
    image_url = None
    if image_file:
        blob = bucket.blob(f"questions/{uuid.uuid4().hex}.jpg")
        blob.upload_from_file(image_file)
        blob.make_public()
        image_url = blob.public_url

    # build choices
    formatted_choices = []
    for letter, (text, is_correct) in (choices or {}).items():
        if text:
            formatted_choices.append({
                "text": text.strip(),
                "is_correct": bool(is_correct)
            })

    # FIRESTORE DOCUMENT
    db.collection("questions").add({
        "question_text": question_text,
        "choices": formatted_choices,
        "correct_letter": next(
            (chr(65 + i) for i, c in enumerate(formatted_choices) if c.get("is_correct")),
            None
        ),
        "image": image_url,
        "difficulty": level,
        "source": source,
        "subject": subject_name,
        "topic": topic_name,
        "board_exams": board_exam_list or [],
        "usage_count": 0,
        "created_at": firestore.SERVER_TIMESTAMP
    })

def extract_and_save_questions(xml_file, subject):

    tree = ET.parse(xml_file)
    root = tree.getroot()

    for question in root.findall('.//question'):

        question_text_element = question.find('questiontext')

        if question_text_element is not None:
            question_text, image_file = extract_question_text_and_image(
                question_text_element,
                subject
            )
        else:
            question_text = ""
            image_file = None

        answers = question.findall('answer')
        if len(answers) < 2:
            continue

        correct_answer = None
        choices = {}
        image_path = None

        for i, answer in enumerate(answers):
            text = strip_tags(answer.find('text').text.strip())
            fraction = int(answer.get('fraction'))

            letter = chr(65 + i)

            if fraction == 100:
                correct_answer = letter

            choices[letter] = (text, fraction == 100)

        if image_file:
            image_path = save_image_locally(image_file)

        save_question_firestore(
            question_text=question_text,
            choices=choices,
            image_file=image_file,
            level="E",
            source="xml_import",
            subject_name=subject,
            topic_name="",
            board_exam_list=None
        )

def extract_question_text_and_image(question_text_element, subject):

    question_text = ""
    image_file = None

    text_element = question_text_element.find('./text')

    if text_element is not None:
        text_content = text_element.text or ""

        match = re.search(r'<p>(.*?)<img', text_content)
        if match:
            question_text = match.group(1).strip()
        else:
            match = re.search(r'<p>(.*?)</p>', text_content)
            if match:
                question_text = match.group(1).strip()

    file_elements = question_text_element.findall('./file')

    for file_element in file_elements:
        image_name = file_element.get('name')

        if image_name.endswith(('.png', '.jpg')):

            file_data = base64.b64decode(file_element.text.strip())

            image_file = InMemoryUploadedFile(
                ContentFile(file_data),
                None,
                image_name,
                'image/jpeg',
                len(file_data),
                None
            )

    return question_text, image_file

def save_image_locally(image_file):

    media_dir = os.path.join("media", "question_images")
    os.makedirs(media_dir, exist_ok=True)

    _, filename = os.path.split(image_file.name)
    path = os.path.join(media_dir, filename)

    with open(path, "wb") as f:
        f.write(image_file.read())

    return path


def upload_xml(request):

    if request.method == "POST" and "xml_file" in request.FILES:

        xml_file = request.FILES["xml_file"]
        subject = request.POST.get("subject")

        extract_and_save_questions(xml_file, subject)

        return HttpResponse("XML uploaded successfully (Firestore)")

    return render(request, "upload_xml.html")

def parse_txt(text, image_files, subject_name, topic_name):

    lines = (text or "").splitlines()
    i = 0
    source = "google.com"

    while i < len(lines):

        line = lines[i].strip()

        if not line:
            i += 1
            continue

        if line.startswith("Source:"):
            source = line.replace("Source:", "").strip()
            i += 1
            continue

        if line.startswith("<Q>"):

            question_text = line.replace("<Q>", "").strip()
            i += 1

            choices = {}
            image_file = None
            difficulty = "E"
            board_exam_list = []

            while i < len(lines) and not lines[i].startswith("<Q>"):

                l = lines[i].strip()

                if l.startswith("Img:"):
                    image_file = l.replace("Img:", "").strip()

                elif l.startswith(">>>"):
                    letter = l[3].strip()[0].upper()
                    text = l[4:].strip()
                    choices[letter] = (text, True)

                elif len(l) > 1 and l[1] == ".":
                    letter = l[0].upper()
                    text = l[2:].strip()
                    choices[letter] = (text, False)

                elif l.upper() in ["VE","E","M","D","VD"]:
                    difficulty = l.upper()

                i += 1

            save_question_firestore(
                question_text,
                choices,
                None,
                difficulty,
                source,
                subject_name,
                topic_name,
                board_exam_list
            )

        i += 1


# -----------------------------------------
# PARSE XLSX
# -----------------------------------------

def save_question_firestore(
    question_text,
    choices,
    image_file,
    level,
    source,
    subject_name,
    topic_name,
    board_exam_list=None
):

    # Upload image to Cloud Storage
    image_url = None
    if image_file:
        blob = bucket.blob(f"questions/{uuid.uuid4().hex}.jpg")
        blob.upload_from_file(image_file)
        blob.make_public()
        image_url = blob.public_url

    formatted_choices = []
    for letter, (text, is_correct) in (choices or {}).items():
        if text:
            formatted_choices.append({
                "text": text.strip(),
                "is_correct": bool(is_correct)
            })

    db.collection("questions").add({
        "question_text": question_text,
        "choices": formatted_choices,
        "image": image_url,
        "difficulty": level,
        "source": source,
        "subject": subject_name,
        "topic": topic_name,
        "board_exams": board_exam_list or [],
        "created_at": firestore.SERVER_TIMESTAMP
    })

def parse_xlsx(df, image_map=None, subject=None, topic=None):

    normalized_cols = {
        "".join(str(c).lower().replace("\xa0", "").split()): c
        for c in df.columns
    }

    def get_col(name):
        key = "".join(name.lower().replace("\xa0", "").split())
        return normalized_cols.get(key)

    q_col = get_col("question")
    a_col = get_col("choicea")
    b_col = get_col("choiceb")
    c_col = get_col("choicec")
    d_col = get_col("choiced")
    e_col = get_col("choicee")
    ans_col = get_col("correctanswer")
    lvl_col = get_col("level")
    img_col = get_col("image")
    src_col = get_col("source")
    be_col = get_col("boardexam")

    for _, row in df.iterrows():

        question_text = str(row.get(q_col, "")).strip() if q_col else ""
        source = str(row.get(src_col, "")).strip() if src_col else "google.com"
        level = str(row.get(lvl_col, "")).strip().upper() if lvl_col else "E"

        image = str(row.get(img_col, "")).strip() if img_col else ""

        if image_map and image in image_map:
            image = image_map[image]

        # choices
        choices_raw = {
            "A": str(row.get(a_col, "")).strip() if a_col else "",
            "B": str(row.get(b_col, "")).strip() if b_col else "",
            "C": str(row.get(c_col, "")).strip() if c_col else "",
            "D": str(row.get(d_col, "")).strip() if d_col else "",
            "E": str(row.get(e_col, "")).strip() if e_col else "",
        }

        correct = str(row.get(ans_col, "")).strip().upper() if ans_col else ""

        choices = {
            k: (v, k == correct)
            for k, v in choices_raw.items()
            if v
        }

        # board exams
        board_exam_list = []
        if be_col and pd.notna(row.get(be_col)):
            raw = str(row[be_col])
            parts = [p.strip().upper() for p in raw.split(",") if p.strip()]
            board_exam_list = parts

        save_question_firestore(
            question_text=question_text,
            choices=choices,
            image_file=image,
            level=level,
            source=source,
            subject_name=subject,
            topic_name=topic,
            board_exam_list=board_exam_list
        )

def upload_file(request):

    if request.method == "POST":

        uploaded_items = request.FILES.getlist("folder_upload")
        subject = request.POST.get("subject")
        topic = request.POST.get("topic")

        main_file = None
        image_map = {}

        # split files
        for f in uploaded_items:
            ext = f.name.lower().split(".")[-1]

            if ext in ["pdf", "docx", "txt", "xlsx"]:
                main_file = f
            elif ext in ["jpg", "jpeg", "png"]:
                image_map[os.path.basename(f.name)] = f

        if not main_file:
            return HttpResponse("No main file found")

        ext = main_file.name.lower().split(".")[-1]

        try:

            if ext == "pdf":
                text = extract_pdf_text(main_file)
                parse_txt(text, image_map, subject, topic)

            elif ext == "docx":
                text = extract_text_from_docx(main_file)
                parse_txt(text, image_map, subject, topic)

            elif ext == "txt":
                text = extract_text_from_txt(main_file)
                parse_txt(text, image_map, subject, topic)

            elif ext == "xlsx":
                df = pd.read_excel(main_file)
                parse_xlsx(df, image_map=image_map, subject=subject, topic=topic)

            else:
                return HttpResponse("Invalid file type")

        except Exception as e:
            return HttpResponse(f"Error: {str(e)}")

        return redirect("question_bank")

    return render(request, "upload_file.html")



####################### FOR UPLOADING AND CHECKING OF ANSWER SHEET (IMAGE) ##############################
from firebase_admin import firestore

db = firestore.client()

def get_exam_id_suggestions(request):
    input_text = request.GET.get('input', '').lower()

    docs = db.collection("answer_keys").stream()

    suggestions = []
    for doc in docs:
        if input_text in doc.id.lower():
            suggestions.append(doc.id)

    return JsonResponse(suggestions, safe=False)

def get_subjects(request):
    docs = db.collection("test_keys").stream()

    subjects = set()

    for doc in docs:
        data = doc.to_dict()
        if "subject" in data:
            subjects.add(data["subject"])

    return JsonResponse({"subjects": list(subjects)})

def get_testkeys_by_subject(request):
    subject = request.GET.get('subject')

    testkeys = []

    if subject:
        docs = db.collection("answer_keys") \
                  .where("subject", "==", subject) \
                  .stream()

        testkeys = [doc.id for doc in docs]

    return JsonResponse({"testkeys": testkeys})

def download_answer_page(request):
    return render(request, 'download_answer_key.html')

def download_exam_results_page(request):
    return render(request, 'download_exam_results.html')

def get_exam_dates_by_board_exam(request):
    board_exam = request.GET.get('board_exam')

    docs = db.collection("test_keys") \
              .where("board_exam", "==", board_exam) \
              .stream()

    dates = set()

    for doc in docs:
        data = doc.to_dict()
        if "exam_date" in data:
            # expected ISO string or timestamp
            try:
                dt = data["exam_date"]
                if isinstance(dt, str):
                    dt = datetime.fromisoformat(dt)

                dates.add(dt.strftime("%B-%Y"))
            except:
                pass

    return JsonResponse({"dates": list(dates)})

def get_subjects_by_board_exam_and_date(request):
    board_exam = request.GET.get('board_exam')
    exam_date = request.GET.get('exam_date')

    month, year = exam_date.split('-')
    month_num = datetime.strptime(month, "%B").month

    docs = db.collection("test_keys") \
              .where("board_exam", "==", board_exam) \
              .stream()

    subjects = set()

    for doc in docs:
        data = doc.to_dict()

        if "exam_date" in data:
            dt = data["exam_date"]

            if isinstance(dt, str):
                dt = datetime.fromisoformat(dt)

            if dt.month == month_num and dt.year == int(year):
                subjects.add(data.get("subject"))

    return JsonResponse({"subjects": list(subjects)})


def view_answer_key(request):
    exam_id = request.GET.get('exam_id')

    doc = db.collection("answer_keys").document(exam_id).get()

    if not doc.exists:
        return JsonResponse({"error": "Not found"})

    return render(request, 'view_answer_key.html', {
        "exam_id": exam_id,
        "answer_key": doc.to_dict().get("answer_key", {})
    })
    
def download_answer_key(request):
    exam_id = request.GET.get('exam_id')

    if not exam_id:
        return JsonResponse({'error': 'Exam ID is required'})

    doc = db.collection("answer_keys").document(exam_id).get()

    if not doc.exists:
        return JsonResponse({'error': 'Answer key not found'})

    data = doc.to_dict()

    file_name = f"answer_key_{exam_id}.txt"

    answer_key_str = (
        f"Board Exam/Course: {data.get('board_exam')}\n"
        f"Subject: {data.get('subject')}\n"
        f"Test Key: {exam_id}\n"
        f"{'-'*40}\n"
        + '\n'.join([f"{k}: {v}" for k, v in data.get("answer_key", {}).items()])
    )

    response = HttpResponse(answer_key_str, content_type='text/plain')
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    return response

# Get distinct board exams
def get_board_exams(request):
    docs = db.collection("board_exams").stream()
    exams = [doc.id for doc in docs]

    return JsonResponse({"board_exams": exams})


# Get distinct subjects by board exam
def get_subjects_by_board_exam(request):
    board_exam = request.GET.get('board_exam')

    docs = db.collection("questions") \
              .where("board_exams", "array_contains", board_exam) \
              .stream()

    subjects = set()

    for doc in docs:
        data = doc.to_dict()
        subjects.add(data.get("subject"))

    return JsonResponse({"subjects": list(subjects)})


# Get distinct topics by subject
def get_topics_by_subject(request):
    subject = request.GET.get('subject')

    docs = db.collection("questions") \
              .where("subject", "==", subject) \
              .stream()

    topics = set()

    for doc in docs:
        data = doc.to_dict()
        topics.add(data.get("topic"))

    return JsonResponse({"topics": list(topics)})


# Get test keys by topic (from AnswerKey)
def get_testkeys_by_topic(request):
    topic = request.GET.get('topic')

    docs = db.collection("questions") \
              .where("topic", "==", topic) \
              .stream()

    subject = None
    for doc in docs:
        subject = doc.to_dict().get("subject")
        break

    testkeys = []

    if subject:
        keys = db.collection("answer_keys") \
                 .where("subject", "==", subject) \
                 .stream()

        testkeys = [doc.id for doc in keys]

    return JsonResponse({"testkeys": testkeys})

from firebase_admin import firestore
from datetime import datetime
from django.http import JsonResponse, HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

db = firestore.client()


def download_exam_results(request):
    subject = request.GET.get('subject')
    exam_date = request.GET.get('exam_date')

    if not subject or not exam_date:
        return JsonResponse({'error': 'Subject and exam date are required'})

    try:
        month, year = exam_date.split('-')
        month_num = datetime.strptime(month, "%B").month
        year = int(year)
    except ValueError:
        return JsonResponse({'error': 'Invalid exam date format'})

    # 🔥 STEP 1: Get matching test_keys from Firestore
    test_docs = db.collection("test_keys") \
        .where("subject", "==", subject) \
        .stream()

    exam_ids = []
    board_exam = None

    for doc in test_docs:
        data = doc.to_dict()

        dt = data.get("exam_date")
        if isinstance(dt, str):
            dt = datetime.fromisoformat(dt)

        if dt and dt.month == month_num and dt.year == year:
            exam_ids.append(doc.id)
            board_exam = data.get("board_exam", board_exam)

    if not exam_ids:
        return JsonResponse({'error': 'No exams found for this subject/date'})

    # 🔥 STEP 2: Get results for those exam_ids
    results_docs = db.collection("results") \
        .where("exam_id", "in", exam_ids[:10])  # Firestore limit safety

    results = list(results_docs.stream())

    if not results:
        return JsonResponse({'message': 'No results yet for this subject/date.'})

    # ================== EXCEL GENERATION ==================
    wb = Workbook()
    ws = wb.active
    ws.title = f"{subject} Results"

    ws["A1"] = "Board Exam:"
    ws["B1"] = board_exam or "-"
    ws["A2"] = "Subject:"
    ws["B2"] = subject
    ws["A3"] = "Exam Date:"
    ws["B3"] = exam_date

    ws.append([])
    ws.append(["Student Name", "Score", "Exam Set"])

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for cell in ws[5]:
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    # 🔥 STEP 3: Write results
    for doc in results:
        data = doc.to_dict()

        ws.append([
            data.get("student_name", ""),
            data.get("score", 0),
            data.get("exam_id", "")
        ])

    # Styling
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = f"Results_{subject}_{exam_date}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response



def image_to_mask(image):
    # Convert image to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # Apply adaptive thresholding
    _, mask = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    
    # Apply morphological operations to clean up the mask
    kernel_size = (2, 2)
    kernel = np.ones(kernel_size, np.uint8)
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
    
    return mask


def upload_answer(request):
    if request.method == 'POST' and request.FILES.get('image'):

        uploaded_image = request.FILES['image']
        exam_id = request.POST.get('exam_id')

        # --- LOAD MODELS ---
        net_original, classes_original = get_original_model()
        net_cropped, classes_cropped = get_cropped_model()

        # --- GET ANSWER KEY (FIRESTORE) ---
        answer_doc = db.collection("answer_keys").document(exam_id).get()

        if not answer_doc.exists:
            return JsonResponse({"error": "Answer key not found"})

        answer_key_data = answer_doc.to_dict()
        subject = answer_key_data.get("subject")
        correct_answers = {
            str(k): v['letter']
            for k, v in answer_key_data.get("answer_key", {}).items()
        }

        # --- GET STUDENT (FIRESTORE instead of Django model) ---
        uid = request.user.id  # or Firebase UID if using Firebase Auth

        student_doc = db.collection("students").document(str(uid)).get()

        if not student_doc.exists:
            return JsonResponse({"error": "Student not found"})

        student = student_doc.to_dict()
        student_id = student.get("student_id")
        course = student.get("course")
        student_name = student.get("last_name", "") + " " + student.get("first_name", "")

        # --- IMAGE DECODE ---
        nparr = np.frombuffer(uploaded_image.read(), np.uint8)
        image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

        # --- MASK ---
        mask_image = image_to_mask(image)

        if mask_image.ndim == 2:
            mask_image = cv2.cvtColor(mask_image, cv2.COLOR_GRAY2BGR)

        reference_point = (0, 0)

        # =========================
        # DETECTION PIPELINE
        # =========================
        boxes_original, _, class_ids_original = detect_objects(
            mask_image, net_original, classes_original
        )

        all_answers = []

        for i, box in enumerate(boxes_original):
            if classes_original[class_ids_original[i]] != 'answer':
                continue

            x, y, w, h = box
            cropped_object = mask_image[y:y+h, x:x+w]

            boxes_cropped, _, class_ids_cropped = detect_objects(
                cropped_object, net_cropped, classes_cropped
            )

            object_dict = sort_objects_by_distance(
                boxes_cropped,
                class_ids_cropped,
                classes_cropped,
                reference_point
            )

            grouped_boxes = group_and_sequence(
                object_dict.values(),
                object_dict.keys()
            )

            seq_num_class_dict = {}
            for seq_num, idx in grouped_boxes.items():
                seq_num_class_dict[seq_num] = classes_cropped[
                    class_ids_cropped[idx - 1]
                ]

            for k in sorted(seq_num_class_dict):
                all_answers.append(seq_num_class_dict[k])

        # =========================
        # SCORING
        # =========================
        score = sum(
            1 for i, a in enumerate(all_answers, start=1)
            if correct_answers.get(str(i)) == a
        )

        total_items = len(correct_answers)

        # =========================
        # CHECK DUPLICATE
        # =========================
        results_ref = db.collection("results") \
            .where("uid", "==", str(uid)) \
            .where("exam_id", "==", exam_id) \
            .stream()

        if any(results_ref):
            return JsonResponse({'warning': 'Answer already uploaded for this exam.'})

        # =========================
        # SAVE RESULT (FIRESTORE)
        # =========================
        db.collection("results").add({
            "uid": str(uid),
            "student_id": student_id,
            "course": course,
            "student_name": student_name,
            "subject": subject,
            "exam_id": exam_id,
            "answer": all_answers,
            "correct_answer": list(correct_answers.values()),
            "score": score,
            "total_items": total_items,
            "is_submitted": True,
            "timestamp": firestore.SERVER_TIMESTAMP
        })

        # cleanup
        del image, mask_image, cropped_object
        gc.collect()

        return JsonResponse({
            "score": score,
            "detected": len(all_answers),
            "total_items": total_items
        })

    return render(request, 'upload_answer.html')



def answer_sheet_view(request):
    if request.method == 'POST':
        form = AnswerSheetForm(request.POST)
        if form.is_valid():
            # TODO: optionally save to Firestore later
            pass
    else:
        form = AnswerSheetForm()

    return render(request, 'answer_sheet.html', {'form': form})

def online_answer_test(request):
    if request.method != 'POST':
        return render(request, 'answer_test_form.html')

    subject = request.POST.get('subject')
    board_exam = request.POST.get('board_exam')
    exam_date_str = timezone.now().strftime("%b%Y")

    set_a_id = f"{board_exam}_{exam_date_str}_{uuid.uuid4().hex[:4]}"
    set_b_id = f"{board_exam}_{exam_date_str}_{uuid.uuid4().hex[:4]}"

    set_a_question_ids = request.POST.getlist('set_a_question_ids[]')
    set_b_question_ids = request.POST.getlist('set_b_question_ids[]')

    # Build questions using your existing helpers
    questions_set_a = get_questions_with_choices(set_a_question_ids)
    questions_set_b = get_questions_with_choices(set_b_question_ids)

    set_a_answer_key = build_answer_key(set_a_question_ids)
    set_b_answer_key = build_answer_key(set_b_question_ids)

    set_a_choice_map = extract_choices_by_letter(questions_set_a)
    set_b_choice_map = extract_choices_by_letter(questions_set_b)

    # -----------------------------
    # FIRESTORE SAVE (testKeys)
    # -----------------------------
    db.collection("testKeys").document(set_a_id).set({
        "set_id": set_a_id,
        "board_exam": board_exam,
        "subject": subject,
        "questions": questions_set_a,
        "choiceA": set_a_choice_map.get("A", []),
        "choiceB": set_a_choice_map.get("B", []),
        "choiceC": set_a_choice_map.get("C", []),
        "choiceD": set_a_choice_map.get("D", []),
        "choiceE": set_a_choice_map.get("E", [])
    })

    db.collection("testKeys").document(set_b_id).set({
        "set_id": set_b_id,
        "board_exam": board_exam,
        "subject": subject,
        "questions": questions_set_b,
        "choiceA": set_b_choice_map.get("A", []),
        "choiceB": set_b_choice_map.get("B", []),
        "choiceC": set_b_choice_map.get("C", []),
        "choiceD": set_b_choice_map.get("D", []),
        "choiceE": set_b_choice_map.get("E", [])
    })

    # -----------------------------
    # FIRESTORE SAVE (answerKeys)
    # -----------------------------
    db.collection("answerKeys").document(set_a_id).set({
        "set_id": set_a_id,
        "board_exam": board_exam,
        "subject": subject,
        "answer_key": set_a_answer_key
    })

    db.collection("answerKeys").document(set_b_id).set({
        "set_id": set_b_id,
        "board_exam": board_exam,
        "subject": subject,
        "answer_key": set_b_answer_key
    })

    # -----------------------------
    # RETURN RESPONSE (no DB query)
    # -----------------------------
    return render(request, 'answer_test.html', {
        'subject': subject,
        'board_exam': board_exam,
        'set_a_questions_choices': questions_set_a,
        'set_b_questions_choices': questions_set_b,
        'set_a_id': set_a_id,
        'set_b_id': set_b_id,
    })



from firebase_admin import firestore
from django.shortcuts import render

db = firestore.client()


def answer_test_preview(request, subject, board_exam, set_a_id, set_b_id):

    # -----------------------
    # GET TEST KEYS
    # -----------------------
    test_a = db.collection("testKeys").document(set_a_id).get().to_dict()
    test_b = db.collection("testKeys").document(set_b_id).get().to_dict()

    # -----------------------
    # GET ANSWER KEYS
    # -----------------------
    answer_a = db.collection("answerKeys").document(set_a_id).get().to_dict()
    answer_b = db.collection("answerKeys").document(set_b_id).get().to_dict()

    return render(request, 'answer_test.html', {
        'subject': subject,
        'board_exam': board_exam,
        'test_a': test_a,
        'test_b': test_b,
        'answer_a': answer_a,
        'answer_b': answer_b,
    })


####################### FOR ANSWERING ONLINE ##############################

import random
import json
from firebase_admin import firestore

db = firestore.client()


def answer_online_exam(request):

    # -----------------------
    # GET ALL TEST KEYS
    # -----------------------
    testkeys = db.collection("testKeys").stream()

    data = {}

    for tk in testkeys:
        tk_data = tk.to_dict()

        board = tk_data.get("board_exam")
        exam_date_raw = tk_data.get("exam_date")  # MUST exist in Firestore
        subject = tk_data.get("subject")

        if not board or not exam_date_raw or not subject:
            continue

        # Convert Firestore timestamp or string
        if hasattr(exam_date_raw, "strftime"):
            month_key = exam_date_raw.strftime("%Y-%m")
            year = exam_date_raw.year
            month = exam_date_raw.month
        else:
            # if stored as string "YYYY-MM-DD"
            try:
                from datetime import datetime
                dt = datetime.fromisoformat(exam_date_raw)
                month_key = dt.strftime("%Y-%m")
                year = dt.year
                month = dt.month
            except:
                continue

        # -----------------------
        # INIT STRUCTURE
        # -----------------------
        data.setdefault(board, {})
        data[board].setdefault(month_key, {})

        if subject not in data[board][month_key]:

            # -----------------------
            # FILTER SAME GROUP
            # -----------------------
            sets_query = db.collection("testKeys") \
                .where("board_exam", "==", board) \
                .where("subject", "==", subject)

            sets_for_subject = [s.to_dict() for s in sets_query.stream()]

            if not sets_for_subject:
                continue

            random_set = random.choice(sets_for_subject)

            data[board][month_key][subject] = {
                "set_id": random_set.get("set_id"),
                "subject_group": random_set.get("subject_group", None)
            }

    return render(
        request,
        "answer_online_exam.html",
        {
            "board_exams": list(data.keys()),
            "exam_data_json": json.dumps(data),
        }
    )

from django.utils.dateparse import parse_datetime

from firebase_admin import firestore
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from django.utils.dateparse import parse_datetime
import random, string, uuid

db = firestore.client()


def exam_form(request, set_id):

    # ===============================
    # GET TEST KEY (FIRESTORE)
    # ===============================
    test_ref = db.collection("testKeys").document(set_id).get()

    if not test_ref.exists:
        return redirect("answer_online_exam")

    test_key = test_ref.to_dict()

    # ===============================
    # GET STUDENT (FIRESTORE or DJANGO HYBRID)
    # ===============================
    student_ref = db.collection("students").document(str(request.user.id)).get()

    if not student_ref.exists:
        return redirect("login")

    student = student_ref.to_dict()

    # ===============================
    # ACCESS CONTROL
    # ===============================
    student_course = (student.get("course") or "").strip().lower()
    exam_code = (test_key.get("board_exam") or "").strip().upper()

    if "civil engineering" in student_course:
        allowed_exam = "CE"
    elif "mechanical engineering" in student_course:
        allowed_exam = "ME"
    elif "electrical engineering" in student_course:
        allowed_exam = "EE"
    elif "electronics engineering" in student_course:
        allowed_exam = "ECE"
    else:
        allowed_exam = None

    if allowed_exam != exam_code:
        messages.error(request, "You are not allowed to access this exam.")
        return redirect("answer_online_exam")

    # ===============================
    # PREVENT DOUBLE SUBMISSION
    # ===============================
    results_query = db.collection("results") \
        .where("user_id", "==", str(request.user.id)) \
        .where("subject", "==", test_key.get("subject")) \
        .stream()

    result = next(results_query, None)

    if result:
        result_data = result.to_dict()
        if result_data.get("is_submitted"):
            messages.error(request, "You have already submitted this subject.")
            return redirect("warning_page")

    # ===============================
    # PREPARE QUESTIONS
    # ===============================
    question_choices = []
    letters = list(string.ascii_uppercase)[:5]

    questions = test_key.get("questions", [])
    choiceA = test_key.get("choiceA", [])
    choiceB = test_key.get("choiceB", [])
    choiceC = test_key.get("choiceC", [])
    choiceD = test_key.get("choiceD", [])
    choiceE = test_key.get("choiceE", [])

    for i, question in enumerate(questions):

        question_text = question.get("question_text") or question.get("question")
        image_url = question.get("image_url") or question.get("image") or ""

        choice_texts = [
            choiceA[i] if i < len(choiceA) else "",
            choiceB[i] if i < len(choiceB) else "",
            choiceC[i] if i < len(choiceC) else "",
            choiceD[i] if i < len(choiceD) else "",
            choiceE[i] if i < len(choiceE) else "",
        ]

        random.shuffle(choice_texts)
        choices = list(zip(letters, choice_texts))

        question_choices.append((question_text, choices, image_url))

    total_items = len(question_choices)

    MAX_ITEMS = 100
    MAX_TIME_SECONDS = 4 * 60 * 60
    total_time_limit = int((total_items / MAX_ITEMS) * MAX_TIME_SECONDS)
    per_question_time_limit = total_time_limit / max(total_items, 1)

    # ===============================
    # POST: SUBMIT EXAM
    # ===============================
    if request.method == "POST":

        if request.session.get("form_submitted"):
            messages.error(request, "You have already submitted this exam.")
            return redirect("warning_page")

        submitted_answers = []

        for i in range(total_items):
            ans = request.POST.get(f"question_{i + 1}")

            if not ans:
                messages.error(request, f"Please select an answer for question {i + 1}")

                return render(request, "exam_form.html", {
                    "test_key": test_key,
                    "question_choices": question_choices,
                    "total_items": total_items,
                    "total_time_limit": total_time_limit,
                    "per_question_time_limit": per_question_time_limit,
                    "start_time": timezone.now().isoformat(),
                })

            submitted_answers.append(ans)

        # ===============================
        # ELAPSED TIME
        # ===============================
        elapsed_time = None
        start_time_str = request.POST.get("start_time")

        if start_time_str:
            start_time = parse_datetime(start_time_str)
            if start_time:
                elapsed_td = timezone.now() - start_time

                h, r = divmod(int(elapsed_td.total_seconds()), 3600)
                m, s = divmod(r, 60)

                elapsed_time = f"{h}hr {m}min {s}sec"

        request.session["form_submitted"] = True

        # ===============================
        # SCORING (FIRESTORE ANSWER KEY)
        # ===============================
        answer_ref = db.collection("answerKeys").document(set_id).get()

        if not answer_ref.exists:
            return redirect("warning_page")

        answer_key = answer_ref.to_dict().get("answer_key", {})

        score = 0
        correct_text_answers = []

        for i, key in enumerate(sorted(answer_key.keys(), key=int)):
            correct_text = answer_key[key].get("text", "")
            correct_text_answers.append(correct_text)

            if i < len(submitted_answers) and submitted_answers[i] == correct_text:
                score += 1

        # ===============================
        # SAVE RESULT (FIRESTORE)
        # ===============================
        result_id = str(uuid.uuid4())

        db.collection("results").document(result_id).set({
            "user_id": str(request.user.id),
            "student_id": student.get("student_id"),
            "course": student.get("course"),
            "student_name": student.get("name"),
            "subject": test_key.get("subject"),
            "exam_id": set_id,
            "answer": submitted_answers,
            "correct_answer": correct_text_answers,
            "score": score,
            "total_items": total_items,
            "is_submitted": True,
            "timestamp": timezone.now().isoformat(),
            "elapsed_time": elapsed_time
        })

        return redirect("result_page", result_id=result_id)

    # ===============================
    # GET REQUEST
    # ===============================
    return render(request, "exam_form.html", {
        "test_key": test_key,
        "question_choices": question_choices,
        "total_items": total_items,
        "total_time_limit": total_time_limit,
        "per_question_time_limit": per_question_time_limit,
        "start_time": timezone.now().isoformat(),
    })


def result_page(request, result_id):

    result_ref = db.collection("results").document(result_id).get()

    if not result_ref.exists:
        return render(request, "result_page.html", {"error": "Result not found"})

    result = result_ref.to_dict()

    percent = 0
    if result.get("total_items"):
        percent = round(
            (result.get("score", 0) / result.get("total_items")) * 100,
            2
        )

    return render(request, "result_page.html", {
        "result": result,
        "percent": percent
    })


def warning_page(request):
    home_student_url = reverse('home_student')  # Assuming 'home_student' is the name of the URL pattern for home_student.html
    return render(request, 'submit_warning.html', {'home_student_url': home_student_url})

def view_results(request):

    user_id = str(request.user.id)

    results_query = db.collection("results") \
        .where("user_id", "==", user_id) \
        .stream()

    results = [r.to_dict() for r in results_query]

    # sort by timestamp (latest first)
    results.sort(key=lambda x: x.get("timestamp", ""), reverse=True)

    return render(request, "view_results.html", {
        "results": results
    })
    
def question_analytics(request):

    questions = db.collection("questions").stream()

    board_exam_counts = {}
    subject_counts = {}
    difficulty_counts = {}

    for q in questions:
        data = q.to_dict()

        board_exams = data.get("board_exams", [])
        subjects = data.get("subjects", [])
        difficulty = data.get("difficulty", "Unknown Difficulty")

        if not board_exams:
            board_exams = ["No Board Exam"]

        if not subjects:
            subjects = ["Unknown Subject"]

        # -------------------
        # BOARD EXAM COUNTS
        # -------------------
        for be in board_exams:
            board_exam_counts[be] = board_exam_counts.get(be, 0) + 1

            for subj in subjects:
                key = (be, subj)
                subject_counts[key] = subject_counts.get(key, 0) + 1

        # -------------------
        # DIFFICULTY COUNTS
        # -------------------
        difficulty_counts[difficulty] = difficulty_counts.get(difficulty, 0) + 1

    # -------------------
    # FORMAT FOR TEMPLATE
    # -------------------
    course_stats = [
        {"board_exam": k, "total": v}
        for k, v in board_exam_counts.items()
    ]

    subject_distribution = [
        {
            "board_exam": be,
            "subject": subj,
            "total_questions": count
        }
        for (be, subj), count in subject_counts.items()
    ]

    context = {
        "course_stats": course_stats,
        "subject_distribution": subject_distribution,

        "course_labels": json.dumps(list(board_exam_counts.keys())),
        "total_questions": json.dumps(list(board_exam_counts.values())),

        "difficulty_labels": json.dumps(list(difficulty_counts.keys())),
        "difficulty_counts": json.dumps(list(difficulty_counts.values())),
    }

    return render(request, "question_analytics.html", context)

def test_analytics(request):
    results_ref = db.collection("results").stream()

    course_results_map = {}

    for doc in results_ref:
        r = doc.to_dict()
        course = r.get("course")

        if not course:
            continue

        course_results_map.setdefault(course, []).append(r)

    course_data = {}

    for course, results in course_results_map.items():

        passed_counts = 0
        failed_counts = 0
        total_score = 0

        question_stats = defaultdict(lambda: {'correct': 0, 'wrong': 0})

        for r in results:
            score = r.get("score", 0)
            total_items = r.get("total_items", 1)

            if score >= 0.6 * total_items:
                passed_counts += 1
            else:
                failed_counts += 1

            total_score += score

            answers = r.get("answer", []) or []
            correct_answers = r.get("correct_answer", []) or []

            for idx, ans in enumerate(answers):
                correct_ans = correct_answers[idx] if idx < len(correct_answers) else None

                if ans == correct_ans:
                    question_stats[idx]['correct'] += 1
                else:
                    question_stats[idx]['wrong'] += 1

        avg_score = total_score / len(results) if results else 0

        question_labels = [f'Q{i+1}' for i in range(len(question_stats))]
        correct_counts = [question_stats[i]['correct'] for i in range(len(question_stats))]
        wrong_counts = [question_stats[i]['wrong'] for i in range(len(question_stats))]

        top_students = sorted(results, key=lambda x: x.get("score", 0), reverse=True)[:10]

        course_data[course] = {
            "passed_counts": passed_counts,
            "failed_counts": failed_counts,
            "avg_score": avg_score,
            "question_labels": question_labels,
            "correct_counts": correct_counts,
            "wrong_counts": wrong_counts,
            "top_students": top_students,
            "passed_json": json.dumps({"passed": passed_counts, "failed": failed_counts}),
            "avg_json": json.dumps({"avg": avg_score}),
            "question_json": json.dumps({
                "labels": question_labels,
                "correct": correct_counts,
                "wrong": wrong_counts
            }),
        }

    return render(request, "test_analytics.html", {
        "course_data": course_data
    })

# ---- Practice: start ----
@require_http_methods(["GET", "POST"])
def practice_start(request):

    student_course_full = request.user.student.course.strip().lower()

    course_mapping = {
        "civil engineering": "CE",
        "mechanical engineering": "ME",
        "electrical engineering": "EE",
        "electronics engineering": "ECE",
    }

    student_course_code = course_mapping.get(student_course_full)

    if not student_course_code:
        messages.error(request, "Your course is not supported for practice exams.")
        return redirect("home")

    subjects = list(BOARD_EXAM_TOPICS.get(student_course_code, {}).keys())

    if request.method == "POST":
        subject_name = request.POST.get("subject")

        try:
            num_items = int(request.POST.get("num_items") or 5)
        except ValueError:
            num_items = 5

        # =========================
        # FIRESTORE QUESTION QUERY
        # =========================
        questions_ref = db.collection("questions") \
            .where("board_exams", "array_contains", student_course_code) \
            .where("subjects", "array_contains", subject_name) \
            .stream()

        qs = []
        for doc in questions_ref:
            q = doc.to_dict()
            q["id"] = doc.id
            qs.append(q)

        if not qs:
            messages.error(request, "No questions found for this subject!")
            return redirect("practice_start")

        available = len(qs)

        if num_items > available:
            messages.error(request, f"You selected {num_items} but only {available} questions exist!")
            return redirect("practice_start")

        random.shuffle(qs)
        chosen = qs[:num_items]

        payload = []
        for q in chosen:
            payload.append({
                "id": q["id"],
                "text": q.get("question_text"),
                "image_name": q.get("image_name"),
                "choices": q.get("choices", []),  # must be stored as list in Firestore
                "correct": q.get("correct_answer"),
            })

        session_id = str(uuid.uuid4())

        request.session[f"practice_{session_id}"] = {
            "board_exam": student_course_code,
            "subject": subject_name,
            "questions": payload,
            "total_items": len(payload),
        }

        request.session.modified = True

        return redirect("practice_take", session_id=session_id)

    return render(request, "practice_start.html", {
        "subjects": subjects,
        "student_course": student_course_code,
    })



# ---- Practice: take (render questions + timer) ----
def practice_take(request, session_id):

    sess_key = f'practice_{session_id}'
    data = request.session.get(sess_key)

    if not data:
        messages.error(request, "Practice session not found or expired.")
        return redirect('practice_start')

    letters = list(string.ascii_uppercase)

    questions_for_client = []

    for qi, q in enumerate(data["questions"], start=1):

        choices = q["choices"].copy()
        random.shuffle(choices)

        for idx, choice in enumerate(choices):
            choice["display_letter"] = letters[idx]

        questions_for_client.append({
            "instance_id": qi,
            "q_id": q["id"],
            "text": q["text"],
            "image_name": q.get("image_name"),
            "choices": choices
        })

    return render(request, "practice_take.html", {
        "session_id": session_id,
        "board_exam": data["board_exam"],
        "questions": questions_for_client,
        "total_items": data["total_items"],
    })



# ---- Practice: submit (grade & analytics) ----
@require_http_methods(["POST"])
def practice_submit(request, session_id):

    sess_key = f'practice_{session_id}'
    data = request.session.get(sess_key)

    if not data:
        messages.error(request, "Practice session not found or expired.")
        return redirect('practice_start')

    questions = data['questions']
    total_items = len(questions)

    results = []
    correct_count = 0
    total_time_elapsed = 0.0

    subject_tracker = {}
    topic_tracker = {}
    difficulty_tracker = {}

    # =========================
    # PROCESS ANSWERS
    # =========================
    for i, q in enumerate(questions, start=1):

        ans = request.POST.get(f'answer_{i}')
        time_spent = float(request.POST.get(f'time_{i}', '0') or 0)

        correct_key = q.get('correct')
        is_correct = (ans == correct_key)

        # ---- NO ORM: fetch from session only (or embed metadata in Firestore later)
        subject = q.get("subject", "Unknown")
        topic = q.get("topic", "Misc")
        difficulty = q.get("difficulty", "Unknown")

        # ================= SUBJECT TRACKER =================
        subject_tracker.setdefault(subject, {"correct": 0, "total": 0, "time": 0})
        subject_tracker[subject]["total"] += 1
        subject_tracker[subject]["time"] += time_spent

        if is_correct:
            subject_tracker[subject]["correct"] += 1

        # ================= TOPIC TRACKER =================
        topic_tracker.setdefault(topic, {"correct": 0, "total": 0, "time": 0, "subject": subject})
        topic_tracker[topic]["total"] += 1
        topic_tracker[topic]["time"] += time_spent

        if is_correct:
            topic_tracker[topic]["correct"] += 1

        # ================= DIFFICULTY TRACKER =================
        difficulty_tracker.setdefault(difficulty, {"correct": 0, "total": 0})
        difficulty_tracker[difficulty]["total"] += 1

        if is_correct:
            difficulty_tracker[difficulty]["correct"] += 1

        if is_correct:
            correct_count += 1

        total_time_elapsed += time_spent

        results.append({
            "index": i,
            "q_id": q["id"],
            "text": q["text"],
            "selected": ans,
            "correct": correct_key,
            "is_correct": is_correct,
            "time_spent": time_spent,
        })

    score = correct_count
    pct = (score / total_items * 100) if total_items else 0

    # =========================
    # FIRESTORE HELPERS
    # =========================
    def update_counter(doc_ref, correct, total, time_sum=None):
        doc = doc_ref.get()
        if doc.exists:
            data_old = doc.to_dict()
        else:
            data_old = {
                "total_items_answered": 0,
                "total_correct": 0,
                "total_attempts": 0,
                "average_time_per_item": 0
            }

        total_items_answered = data_old.get("total_items_answered", 0) + total
        total_correct = data_old.get("total_correct", 0) + correct
        total_attempts = data_old.get("total_attempts", 0) + 1

        prev_avg_time = data_old.get("average_time_per_item", 0)
        prev_total = data_old.get("total_items_answered", 0)

        new_time = (prev_avg_time * prev_total) + (time_sum or 0)
        avg_time = new_time / total_items_answered if total_items_answered else 0

        doc_ref.set({
            "total_items_answered": total_items_answered,
            "total_correct": total_correct,
            "total_attempts": total_attempts,
            "average_time_per_item": avg_time
        }, merge=True)

    # =========================
    # UPDATE SUBJECT ANALYTICS
    # =========================
    for subject, stats in subject_tracker.items():
        doc_ref = db.collection("subject_analytics") \
            .document(f"{request.user.id}_{subject}_{data['board_exam']}")

        update_counter(doc_ref, stats["correct"], stats["total"], stats["time"])

    # =========================
    # UPDATE TOPIC ANALYTICS
    # =========================
    for topic, stats in topic_tracker.items():
        doc_ref = db.collection("topic_analytics") \
            .document(f"{request.user.id}_{topic}")

        update_counter(doc_ref, stats["correct"], stats["total"], stats["time"])

    # =========================
    # UPDATE DIFFICULTY ANALYTICS
    # =========================
    for difficulty, stats in difficulty_tracker.items():
        doc_ref = db.collection("difficulty_analytics") \
            .document(f"{request.user.id}_{difficulty}_{data['board_exam']}")

        doc_ref.set({
            "total_items_answered": firestore.Increment(stats["total"]),
            "total_correct": firestore.Increment(stats["correct"])
        }, merge=True)

    # =========================
    # SAVE PRACTICE RESULT
    # =========================
    db.collection("practice_results").add({
        "session_id": session_id,
        "user_id": request.user.id,
        "board_exam": data["board_exam"],
        "total_items": total_items,
        "score": score,
        "percent": pct,
        "total_time": total_time_elapsed,
        "answers": results,
        "created_at": firestore.SERVER_TIMESTAMP
    })

    # =========================
    # STORE SESSION RESULT
    # =========================
    request.session[f'practice_result_{session_id}'] = {
        "score": score,
        "total_items": total_items,
        "percent": pct,
        "results": results,
        "total_time": total_time_elapsed,
        "board_exam": data["board_exam"],
        "created_at": timezone.now().isoformat(),
    }

    request.session.modified = True

    return redirect('practice_result_page', session_id=session_id)


def practice_result_page(request, session_id):

    doc = db.collection("practice_results") \
        .document(session_id) \
        .get()

    if not doc.exists:
        messages.error(request, "No practice results found for that session.")
        return redirect('practice_start')

    res = doc.to_dict()

    return render(request, 'practice_result.html', {
        'res': res
    })

from django.core.serializers.json import DjangoJSONEncoder

@login_required
def analytics_dashboard(request):

    user_id = request.user.id

    results_ref = db.collection("practice_results") \
        .where("user_id", "==", user_id) \
        .stream()

    subject_data = {}
    topic_data = {}
    difficulty_data = {}

    for doc in results_ref:
        res = doc.to_dict()

        board_exam = res.get("board_exam")
        answers = res.get("answers", [])

        for ans in answers:

            q_id = ans.get("q_id")
            selected = ans.get("selected")
            correct = ans.get("correct")
            time_spent = ans.get("time_spent", 0.0)

            # =========================
            # QUESTION FETCH (FIRESTORE)
            # =========================
            q_doc = db.collection("questions").document(q_id).get()
            if not q_doc.exists:
                continue

            q = q_doc.to_dict()

            subjects = q.get("subjects", [])
            topic = q.get("topic", "Misc")
            difficulty = q.get("difficulty", "Unknown")

            subject = "Unknown"
            for subj in subjects:
                if subj in BOARD_EXAM_TOPICS.get(board_exam, {}):
                    subject = subj
                    break

            # ================= SUBJECT =================
            key = (subject, board_exam)
            subject_data.setdefault(key, {
                "total_items_answered": 0,
                "total_correct": 0,
                "total_time": 0.0,
            })

            subject_data[key]["total_items_answered"] += 1
            subject_data[key]["total_correct"] += int(selected == correct)
            subject_data[key]["total_time"] += time_spent

            # ================= TOPIC =================
            key_topic = (topic, subject)
            topic_data.setdefault(key_topic, {
                "total_items_answered": 0,
                "total_correct": 0,
                "total_time": 0.0,
            })

            topic_data[key_topic]["total_items_answered"] += 1
            topic_data[key_topic]["total_correct"] += int(selected == correct)
            topic_data[key_topic]["total_time"] += time_spent

            # ================= DIFFICULTY =================
            key_diff = (difficulty, board_exam)
            difficulty_data.setdefault(key_diff, {
                "total_items_answered": 0,
                "total_correct": 0,
            })

            difficulty_data[key_diff]["total_items_answered"] += 1
            difficulty_data[key_diff]["total_correct"] += int(selected == correct)

    # =========================
    # BUILD SUBJECT LIST
    # =========================
    subject_list = []
    for (subj, be), v in subject_data.items():

        avg_time = v["total_time"] / v["total_items_answered"] if v["total_items_answered"] else 0
        acc = (v["total_correct"] / v["total_items_answered"] * 100) if v["total_items_answered"] else 0

        subject_list.append({
            "subject": subj,
            "board_exam": be,
            "total_items_answered": v["total_items_answered"],
            "total_correct": v["total_correct"],
            "average_time_per_item": round(avg_time, 2),
            "accuracy": round(acc, 2),
        })

    # =========================
    # BUILD TOPIC LIST
    # =========================
    topic_list = []
    for (topic, subj), v in topic_data.items():

        avg_time = v["total_time"] / v["total_items_answered"] if v["total_items_answered"] else 0
        acc = (v["total_correct"] / v["total_items_answered"] * 100) if v["total_items_answered"] else 0

        topic_list.append({
            "subject": subj,
            "topic": topic,
            "total_items_answered": v["total_items_answered"],
            "total_correct": v["total_correct"],
            "average_time_per_item": round(avg_time, 2),
            "accuracy": round(acc, 2),
        })

    # =========================
    # BUILD DIFFICULTY LIST
    # =========================
    difficulty_list = []
    for (diff, be), v in difficulty_data.items():

        acc = (v["total_correct"] / v["total_items_answered"] * 100) if v["total_items_answered"] else 0

        difficulty_list.append({
            "difficulty": diff,
            "board_exam": be,
            "total_items_answered": v["total_items_answered"],
            "total_correct": v["total_correct"],
            "accuracy": round(acc, 2),
        })

    return render(request, "analytics_dashboard.html", {
        "subject_analytics": json.dumps(subject_list),
        "topic_analytics": json.dumps(topic_list),
        "difficulty_analytics": json.dumps(difficulty_list),
        "ai_suggestions": "",
    })
